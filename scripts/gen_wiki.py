# -*- coding: utf-8 -*-
"""Generate the Economics Academy LLM Wiki (markdown) from the course-map workbook.
Single source of truth = the xlsx. Re-run whenever the spine changes."""
import openpyxl, re, datetime

WB = "outputs/Economics-Academy-Course-Map-v3.xlsx"
OUT = "outputs/economics-academy-llm-wiki.md"
wb = openpyxl.load_workbook(WB)

def rows(ws, start=1):
    for r in range(start, ws.max_row + 1):
        yield [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

L = []
A = L.append

# ============================================================ header / usage
A("# STOIC EDU · Economics Academy — LLM Wiki (v3)")
A("")
A(f"_Auto-generated from `Economics-Academy-Course-Map-v3.xlsx` on {datetime.date.today().isoformat()}. "
  "The workbook is the single source of truth; never edit this wiki directly — edit the workbook and regenerate._")
A("")
A("## How to use this wiki (instructions for the LLM)")
A("")
A("You are creating teaching materials (slides, games, worksheets, questions, notes) for Stoic Edu's "
  "Pre-IGCSE / Pre-IB Economics Academy. Before producing anything, obey the PRIME DIRECTIVES below, "
  "then locate the target lesson in **§6 Lesson Specs** and follow its spec exactly. "
  "Lesson IDs use the form `M{module}L{lesson}` (e.g. M1L3 = Module 1, Lesson 3).")
A("")

# ============================================================ prime directives
A("## §0 PRIME DIRECTIVES (always apply)")
A("")
A("1. **Three layers, never mixed.** One conceptual SPINE -> two thinking TIERS -> a swappable EXAM ADAPTER. "
  "Tier 1 'Draw the Map' (ages 10-12, sold as Pre-IGCSE): concepts discovered through games, simulations, stories — "
  "no exam content, no heavy calculation. Tier 2 'Read the Map' (ages 13-15, sold as Pre-IB): the same topics promoted "
  "into definitions, diagrams, calculations, evaluative writing. Exam technique exists ONLY in the Exam Adapter Slot "
  "(10-15 min at lesson close, Tier 2 only), with content taken from §7 — never invented per lesson.")
A("2. **Same topic, different verbs.** Tier 1 feels it (Play · Discover · Feel); Tier 2 judges it (Judge · Evaluate · Connect). "
  "Never write a Tier 1 material that defines before playing; never write a Tier 2 material without the promotion to formal treatment.")
A("3. **Tiers are entry points, not a sequence.** A 13-15 year-old IB-bound beginner enters directly at Tier 2; "
  "Tier 1 is not a prerequisite. Never imply Pre-IGCSE must precede Pre-IB.")
A("4. **Honour the content flags** in every lesson spec: bullets are in PRIORITY ORDER (top = most important); "
  "`[ENRICH]` = beyond one or both syllabuses — optional, clearly labelled, first thing to cut when time is short; "
  "`[IB SL CORE]` = core for IB-bound students only, enrichment for IGCSE-bound; "
  "`[LOW]` = minor — keep light, do not over-invest, never expand; `(!)` = scope note that constrains what you may produce.")
A("5. **Nothing outside the current syllabuses is ever taught as core.** Hard boundaries: "
  "XED (in neither syllabus); deflation, fixed/managed exchange rates, market-failure D/S diagrams, the AD-AS model "
  "(all absent from IGCSE 0455 2027-29); the formal comparative-advantage model and the multiplier (IB HL only). "
  "These may appear only as labelled enrichment.")
A("6. **Single-point updates.** When adding a case, news hook or game, attach it to the SPINE content of one lesson "
  "so it serves both tiers; do not fork tier-specific content unless the spec's tier columns already differ.")
A("7. **All questions must be original.** Never reproduce past-paper items from either board.")
A("")

# ============================================================ overview / commitments
A("## §1 Purpose, Commitments & Architecture")
A("")
ws = wb["Overview"]
for row in rows(ws, 4):
    k, v = row[0], row[1]
    if k and v:
        A(f"**{k}.** {v}")
        A("")

# ============================================================ structure
A("## §2 Curriculum Structure")
A("")
ws = wb["Structure"]
hdr = None
for row in rows(ws, 3):
    cells = [c for c in row if c]
    if not cells:
        continue
    if hdr is None:
        hdr = row[:5]
        A("| " + " | ".join(str(h) for h in hdr) + " |")
        A("|" + "---|" * len(hdr))
    else:
        A("| " + " | ".join(str(c) if c else "" for c in row[:5]) + " |")
A("")

# ============================================================ key concepts
A("## §3 The Nine IB Key Concepts (the spine's lenses)")
A("")
A("PP = primary lens for that module; P = revisited.")
A("")
ws = wb["Key-Concept Map"]
A("| Key concept | Micro (M1) | Macro (M2) | Global (M3) |")
A("|---|---|---|---|")
for row in rows(ws, 4):
    if row[0]:
        A(f"| {row[0]} | {row[1]} | {row[2]} | {row[3]} |")
A("")

# ============================================================ lesson template
A("## §4 Standard Lesson Template (12 sections, layer-tagged)")
A("")
A("Every lesson material you produce must slot into one of these sections. The Layer column tells you the editing rule: "
  "SPINE = write once, serves both tiers; TIER SPLIT = produce both tier versions; ADAPTER = content comes from §7 only.")
A("")
ws = wb["Lesson Template"]
A("| # | Section | Layer | Tier 1 delivery | Tier 2 delivery |")
A("|---|---|---|---|---|")
for row in rows(ws, 4):
    if row[0] and str(row[0]).isdigit():
        cells = [str(c).replace("\n", " ") if c else "" for c in row[:5]]
        A("| " + " | ".join(cells) + " |")
# footers (merged notes)
for row in rows(ws, 4):
    if row[0] and not str(row[0]).isdigit():
        A("")
        A(f"> {row[0]}")
A("")

# ============================================================ lesson specs
A("## §5 Module Footnotes (syllabus alignment per module)")
A("")
A("## §6 LESSON SPECS — the 18 lessons")
A("")
mod_sheets = ["M1 Microeconomics", "M2 Macroeconomics", "M3 Global Economics"]
headers = ["#", "Title", "Core Concepts", "Essential Questions", "Core content",
           "Key diagrams", "Exam-skill focus (Tier 2)", "Tier 1 track (ages 10-12)", "Tier 2 track (ages 13-15)"]
for mi, name in enumerate(mod_sheets, 1):
    ws = wb[name]
    A(f"### Module {mi} — {name.split(' ', 1)[1]}")
    A("")
    footer = None
    for row in rows(ws, 4):
        num = row[0]
        if num and str(num).strip().isdigit():
            lid = f"M{mi}L{num}"
            title = row[1]
            A(f"#### {lid} · {title}")
            A("")
            fields = [("Core Concepts", row[2]), ("Essential Questions", row[3]),
                      ("Core content (priority-ordered)", row[4]), ("Key diagrams", row[5]),
                      ("Exam-skill focus (Tier 2)", row[6]),
                      ("Tier 1 delivery — Draw the Map", row[7]),
                      ("Tier 2 delivery — Read the Map", row[8])]
            for fname, val in fields:
                if not val:
                    continue
                A(f"**{fname}:**")
                for line in str(val).split("\n"):
                    line = line.strip()
                    if not line:
                        A("")
                    elif line.startswith("\u2022"):
                        A(f"- {line[1:].strip()}")
                    elif re.match(r"^\d+\.", line):
                        A(f"{line}")
                    elif line.startswith("[") or line.startswith("(!)") or line.startswith("KC:"):
                        A(f"- {line}")
                    else:
                        A(line)
                A("")
        elif row[0] and "alignment" in str(row[0]):
            footer = row[0]
    if footer:
        A(f"> **Module {mi} syllabus alignment:** {footer}")
        A("")

# ============================================================ drip map
A("## §7 EXAM ADAPTER DRIP MAP")
A("")
A("The adapter is drip-fed inside the 18 lessons — never a separate module. When writing an Exam Adapter Slot, "
  "take THIS lesson's row: shared skill first; then the IGCSE element for IGCSE-bound students, the IB element for "
  "IB-bound; before M2 L4 both practise the shared element together.")
A("")
ws = wb["Exam Skills Progression"]
A("| Stage | Lesson | Shared skill (both boards) | Command terms | IGCSE adapter element | IB adapter element |")
A("|---|---|---|---|---|---|")
stage = ""
for row in rows(ws, 4):
    lesson = row[1]
    if lesson and re.match(r"M\d L\d", str(lesson)):
        s = row[0] if row[0] else stage
        if row[0]:
            stage = row[0]
        cells = [str(c).replace("\n", " ") if c else "" for c in [stage, *row[1:6]]]
        A("| " + " | ".join(cells) + " |")
A("")

# ============================================================ bridge
A("## §8 THE IGCSE -> IB BRIDGE (concept promotion map)")
A("")
A("Use this when a material should show progression, motivate a concept's future payoff, or serve parent-facing "
  "explanation. Each row: what the concept is at IGCSE -> what it becomes at IB -> the representative "
  "read-the-world question that frames it.")
A("")
ws = wb["IGCSE to IB Bridge"]
current_section = None
for row in rows(ws, 5):
    first = row[0]
    if first and str(first).startswith("|"):
        current_section = str(first).lstrip("| ").strip()
        A(f"### {current_section}")
        A("")
        A("| IGCSE concept | IB expansion | Added lens | Representative IB question | Level | Lessons |")
        A("|---|---|---|---|---|---|")
    elif first and current_section and row[1]:
        cells = [str(c).replace("\n", " ") if c else "" for c in row[:6]]
        A("| " + " | ".join(cells) + " |")
A("")
# key concepts strip + verification note
for row in rows(ws, 5):
    v = row[0]
    if v and ("Where every bridge leads" in str(v) or "Verification note" in str(v)):
        A(f"> {v}")
        A("")

# ============================================================ verification facts
A("## §9 SYLLABUS VERIFICATION FACTS (checked against official documents, July 2026)")
A("")
ws = wb["Verification Notes"]
for row in rows(ws, 4):
    if row[0] and row[1]:
        A(f"- **{row[0]} {row[1]}** — {row[2]} _Action: {row[3]}_")
A("")

# ============================================================ assessment styles
A("## §10 Assessment Styles Practised (Tier 2)")
A("")
ws = wb["Assessment Styles"]
A("| Question type | Modelled on | First appears |")
A("|---|---|---|")
for row in rows(ws, 4):
    if row[0]:
        A(f"| {row[0]} | {row[1]} | {row[2]} |")
A("")

# ============================================================ glossary
A("## §11 Internal Glossary")
A("")
gloss = [
 ("Spine", "The single shared conceptual core of all 18 lessons; the only place content updates happen."),
 ("Tier 1 / Draw the Map", "Ages 10-12 delivery layer (public label: Pre-IGCSE). Verbs: play, discover, feel."),
 ("Tier 2 / Read the Map", "Ages 13-15 delivery layer (public label: Pre-IB). Verbs: judge, evaluate, connect."),
 ("Exam adapter", "The swappable board-specific top layer (IGCSE: 3 AOs; IB: 4 AOs + command terms + RWE). "
                  "Fitted 1-2 years before the exam; drip-fed via §7."),
 ("Adapter slot", "The 10-15 minute exam-technique segment closing every Tier 2 lesson (Lesson Template section 9)."),
 ("Drip map", "§7 — the per-lesson schedule of which adapter element loads when."),
 ("RWE", "Real-world example — required evidence in IB Paper 1(b) essays; the RWE bank starts collecting from M3."),
 ("Promotion", "The upgrade of a concept from Tier 1 experience to Tier 2 formal treatment, and from IGCSE knowledge to IB lens."),
 ("[ENRICH] / [IB SL CORE] / [LOW] / (!)", "The content boundary flags — see Prime Directive 4."),
 ("Levels-of-response", "The IGCSE Paper 2 (d)-question banded marking style (8 marks, two-sided + conclusion)."),
 ("Command terms", "The board-defined instruction verbs (define, explain, discuss...) that determine answer structure."),
]
for term, desc in gloss:
    A(f"- **{term}** — {desc}")
A("")
A("---")
A("_End of wiki. If a request conflicts with a Prime Directive, the directive wins; say so and propose a compliant alternative._")

open(OUT, "w", encoding="utf-8", newline="\n").write("\n".join(L))
print("wiki written:", OUT, "|", len("\n".join(L)), "chars")
