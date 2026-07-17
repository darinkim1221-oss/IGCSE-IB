# -*- coding: utf-8 -*-
"""Build updated Economics Academy Course Map:
- verified against Cambridge IGCSE 0455 (2027-2029) & IB DP Economics guide (first assessment 2022)
- Core Concepts filled per lesson
- Essential Questions column added (2-3 per lesson)
- Level differentiation: Pre-IGCSE (10-12) vs Pre-IB (13-15)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "0D2D52"; GOLD = "C8880A"; CREAM = "FAF7F2"
LIGHT_BLUE = "DCE9F7"   # Pre-IGCSE column tint
LIGHT_GREEN = "E3EFE0"  # Pre-IB column tint
LIGHT_GOLD = "FBEFD9"   # essential questions tint
GREY = "F2F2F2"
FIX = "FFF2CC"  # corrected cells highlight

F = lambda **kw: Font(name="Arial", **kw)
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="center", horizontal="center")

wb = openpyxl.Workbook()
wb.remove(wb.active)

def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = F(size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = F(size=10, italic=True, color=NAVY)
    c.fill = PatternFill("solid", fgColor=CREAM)
    c.alignment = WRAP
    ws.row_dimensions[2].height = 28

def header_row(ws, row, headers, fills=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F(size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=(fills[i-1] if fills else NAVY))
        c.alignment = WRAP_C
        c.border = BORDER
    ws.row_dimensions[row].height = 30

def put(ws, row, values, fills=None, bold_first=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = F(size=10, bold=(bold_first and i == 1))
        c.alignment = WRAP
        c.border = BORDER
        if fills and fills[i-1]:
            c.fill = PatternFill("solid", fgColor=fills[i-1])

# ============================================================ IGCSE to IB Bridge (first tab)
ws = wb.create_sheet("IGCSE to IB Bridge")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
c = ws.cell(row=1, column=1, value="The Bridge  |  IGCSE is where students draw the map of economics — IB is where they take that map and read the world with it")
c.font = F(size=15, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(vertical="center", wrap_text=True); ws.row_dimensions[1].height = 34
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
c = ws.cell(row=2, column=1, value="The same concept does not simply recur — it is promoted from knowledge to lens. What the student defines and draws at IGCSE becomes, at IB, an instrument for judgement and evaluation. The final column shows where each bridge is built across this 18-lesson course map. Verified against Cambridge IGCSE 0455 (2027-29) and the IB DP Economics Guide (2022); corrections and additions are marked [FIX] / [ADDED].")
c.font = F(size=10, italic=True, color=NAVY); c.fill = PatternFill("solid", fgColor=CREAM)
c.alignment = WRAP; ws.row_dimensions[2].height = 46

# journey strip
c = ws.cell(row=3, column=1, value="IGCSE — Drawing the Map")
c.font = F(size=10, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="4A7AB5"); c.alignment = WRAP_C
ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)
c = ws.cell(row=3, column=2, value="IB — Reading the World")
c.font = F(size=10, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="5E8C61"); c.alignment = WRAP_C
c = ws.cell(row=3, column=5, value="Level"); c.font = F(size=10, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = WRAP_C
c = ws.cell(row=3, column=6, value="Course map"); c.font = F(size=10, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=GOLD); c.alignment = WRAP_C
ws.row_dimensions[3].height = 24
header_row(ws, 4, ["IGCSE concept", "How it expands at IB", "The thinking IB adds (concept / lens)", "A representative IB question (read-the-world)", "Level", "Lesson mapping"],
           fills=["4A7AB5", "5E8C61", "5E8C61", "5E8C61", NAVY, GOLD])

IGF="E7EEF7"; IBF="E9F2E7"; QF="E9F2E7"; LVF="F2F2F2"; MAPF=LIGHT_GOLD
def bridge_rows(section_en, rows, start_r, tag=""):
    r = start_r
    for col in range(1, 7):
        c = ws.cell(row=r, column=col)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = Border(left=thin, right=thin, top=Side(style="medium", color=GOLD),
                          bottom=Side(style="medium", color=GOLD))
    c = ws.cell(row=r, column=1)
    c.value = "|  " + section_en + ("      |   " + tag if tag else "")
    c.font = F(size=11, bold=True, color="FFFFFF")
    c.alignment = Alignment(vertical="center", wrap_text=False)
    ws.row_dimensions[r].height = 22
    r += 1
    for row in rows:
        put(ws, r, list(row), fills=[IGF, IBF, IBF, QF, LVF, MAPF])
        ws.cell(row=r, column=1).font = F(size=10, bold=True)
        ws.cell(row=r, column=5).alignment = WRAP_C
        ws.row_dimensions[r].height = 46
        r += 1
    return r

r = 5
r = bridge_rows("MICRO FOUNDATIONS", tag="IGCSE Topics 1-2 (micro) to IB Units 1-2 (micro)", rows=[
 ("Scarcity", "Resource allocation and the comparison of economic systems", "Efficiency, Equity, Sustainability", "Which economic system allocates scarce resources most efficiently?", "SL", "M1 L1"),
 ("Choice & Opportunity Cost", "Rational decision-making, and the critique of rationality itself (behavioural economics) [FIX: the critique of rationality is HL 2.4]", "Trade-offs, Incentives", "What opportunity cost lies hidden in a government's choices?", "SL (+HL)", "M1 L1"),
 ("PPC [ADDED]", "Promoted into the language of growth analysis: actual vs potential growth, linked to LRAS", "Change, Efficiency", "What is the difference between a movement along the curve and a shift of the curve itself?", "SL", "M1 L1 -> M2 L4"),
 ("Factors of Production", "Productivity and economic growth", "Human capital, Capital deepening", "How does a rise in productivity affect long-run growth?", "SL", "M1 L1 -> M2 L5"),
 ("Demand & Supply", "Consumer and producer behaviour — and the testing of the assumption that this behaviour is 'rational'", "Market equilibrium, Welfare", "Does the market always produce an efficient outcome?", "SL (+HL 2.4)", "M1 L2"),
 ("Price Changes", "Consumer and producer surplus", "Welfare analysis", "When a price rises, who gains and who loses?", "SL", "M1 L2"),
 ("Market Equilibrium", "Allocative efficiency (MB=MC) and the signalling function of price", "Allocative efficiency", "Is the equilibrium price also the socially optimal one?", "SL", "M1 L2"),
 ("Elasticity (PED/PES)", "Government policy analysis and tax incidence [FIX: IB adds YED; XED was removed in the 2022 guide; incidence calculation is HL]", "Tax incidence, Revenue, Welfare", "How does elasticity determine the effect of a tax policy?", "SL (+HL calc.)", "M1 L3"),
], start_r=r)
r = bridge_rows("DECISION-MAKERS", tag="IGCSE Topic 3 (micro) -> at IB it divides: Firms stay micro (Unit 2), Money & labour are promoted to macro (Unit 3)", rows=[
 ("Money & Banking [ADDED]", "The transmission channel of monetary policy — the functions of money become the central bank's lever", "Interest-rate transmission, Central bank credibility", "How far can a central bank actually steer an economy?", "SL", "M2 L1 -> M2 L5"),
 ("Households & Workers [ADDED]", "Labour treated as a market: the minimum wage as a price floor, wage gaps opening into the analysis of inequality", "Equity, Incentives", "Why is some labour paid more than other labour — and is that fair?", "SL (+HL 2.12)", "M2 L1 -> M3 L4"),
 ("Firms, Costs & Market Types [ADDED]", "HL market power: the spectrum from perfect competition to monopoly, and the profit-maximisation model", "Profit maximisation, Efficiency vs Power", "When does a single seller become a problem?", "HL", "M1 L5"),
], start_r=r)
r = bridge_rows("INTERVENTION & FAILURE", tag="IGCSE Topic 2 (micro) to IB Unit 2 (micro)", rows=[
 ("Externalities", "Market failure [FIX: the IB 2022 guide adds emphasis on common pool resources and sustainability]", "Social cost / Social benefit, Sustainability", "Why do markets fail?", "SL", "M1 L4"),
 ("Public Goods", "Government intervention", "Free-rider problem", "How far should a government intervene?", "SL", "M1 L4"),
 ("Merit & Demerit Goods", "[FIX] The IB 2022 guide reframes these terms around externalities and information problems; extended into behavioural economics (HL 2.4) and asymmetric information (HL 2.10)", "Information failure", "Are consumers always rational?", "SL (+HL)", "M1 L4 (+L2, L5 enrichment)"),
 ("Taxes", "Fiscal policy and redistribution (IB 3.4 progressive taxation)", "Equity vs Efficiency", "How far can taxation correct market failure?", "SL", "M1 L4 -> M2 L5 -> M3 L4"),
 ("Subsidies", "Welfare improvement — and the risk of government failure", "Government failure", "Is a subsidy always good policy?", "SL", "M1 L4"),
 ("Price Ceiling / Floor", "Price controls: their intent and their unintended consequences", "Shortage, Surplus, Black markets", "Do price controls produce their intended result?", "SL", "M1 L4"),
], start_r=r)
r = bridge_rows("MACRO", tag="IGCSE Topic 4 (macro) to IB Unit 3 (macro)", rows=[
 ("GDP", "Living standards — the limits of the measure and the alternatives to it", "Well-being, Inequality", "Does GDP genuinely measure quality of life?", "SL", "M2 L2"),
 ("Inflation", "Macroeconomic objectives, analysed through AD-AS (demand-pull vs cost-push)", "AD-AS analysis (+HL Phillips curve)", "What is the best policy for restraining inflation?", "SL (+HL)", "M2 L4"),
 ("Unemployment", "Labour-market analysis — causes by type and the policy responses to each", "Structural vs Cyclical (+HL natural rate)", "Is unemployment a problem the government must solve?", "SL (+HL)", "M2 L4"),
 ("Economic Growth", "Long-run development", "Productivity, Sustainability", "Are growth and the environment reconcilable?", "SL", "M2 L4"),
 ("Fiscal Policy", "Demand management [FIX: the multiplier effect is HL-only]", "Multiplier effect (HL)", "How effective is fiscal policy?", "SL (+HL)", "M2 L5"),
 ("Monetary Policy", "Interest-rate transmission", "Inflation targeting", "Does raising interest rates always lower inflation?", "SL", "M2 L5"),
], start_r=r)
r = bridge_rows("GLOBAL", tag="IGCSE Topic 6 (international) to IB Unit 4 (The Global Economy)", rows=[
 ("International Trade", "Comparative advantage [FIX: the formal comparative-advantage model is HL-only; the gains from trade themselves are SL]", "Globalisation, Interdependence", "Does free trade benefit everyone?", "SL (+HL model)", "M3 L1"),
 ("Protectionism", "Trade-policy evaluation", "Winners & Losers", "When can protectionism be justified?", "SL", "M3 L2"),
 ("Exchange Rates", "Linked to the balance of payments [FIX: fixed/managed rates are HL; IGCSE 2027-29 covers floating only]", "Competitiveness", "How do exchange-rate movements affect the whole economy?", "SL (+HL)", "M3 L3"),
 ("Globalisation & MNCs [ADDED]", "The ladder of economic integration (FTA -> customs union -> common market -> monetary union)", "Interdependence", "How does integration trade sovereignty for prosperity?", "SL (+HL deeper)", "M3 L5"),
], start_r=r)
r = bridge_rows("EQUITY & DEVELOPMENT", tag="IGCSE Topic 5 (development) to IB Unit 3.4 (inequality, SL core) + Unit 4 (development)", rows=[
 ("Poverty & Inequality [ADDED]", "IB 3.4 (SL CORE): Lorenz curve, Gini coefficient, evaluation of redistribution policy", "Equity — the destination of this bridge among the nine key concepts", "How unequal is too unequal?", "SL core", "M3 L4"),
 ("Population [ADDED]", "A background variable of development economics (population structure driving growth and welfare pressure)", "Change, Economic well-being", "Is a population a burden or an engine?", "IGCSE core / IB context", "M3 L4"),
 ("Development", "Development economics — distinguishing growth from development, evaluating strategies", "Poverty, HDI, Inequality", "Does economic growth necessarily mean development?", "SL", "M3 L5"),
], start_r=r)

# nine key concepts destination strip
ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=6)
c = ws.cell(row=r+1, column=1, value="Where every bridge leads — the nine IB Key Concepts (every IGCSE concept ultimately converges on these nine lenses): Scarcity · Choice · Efficiency · Equity · Economic well-being · Sustainability · Change · Interdependence · Intervention")
c.font = F(size=10, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=GOLD)
c.alignment = WRAP; ws.row_dimensions[r+1].height = 30
ws.merge_cells(start_row=r+2, start_column=1, end_row=r+2, end_column=6)
c = ws.cell(row=r+2, column=1, value="Verification note: of the original table, 16 rows were retained; 7 rows were corrected [FIX] (explicit HL boundaries drawn around the comparative-advantage model, the multiplier effect, tax-incidence calculation, behavioural economics, asymmetric information and fixed exchange rates; the IB 2022 reframing of merit/demerit goods; the removal of XED and addition of YED in elasticity; and the strengthened treatment of common pool resources under externalities); and 7 rows were added [ADDED] (PPC, Money & Banking, Households & Workers, Firms & Market Types, Globalisation & MNCs, Poverty & Inequality, Population — all present in both boards' syllabuses but absent from the original table).")
c.font = F(size=9, italic=True, color=NAVY); c.fill = PatternFill("solid", fgColor=CREAM)
c.alignment = WRAP; ws.row_dimensions[r+2].height = 46

for col, w in zip("ABCDEF", [28, 48, 33, 44, 15, 21]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ============================================================ Overview
ws = wb.create_sheet("Overview")
title_block(ws, "Pre-IGCSE / Pre-IB Economics Academy — Foundation Curriculum (v3, 18 lessons, audit-complete)",
            "Complete curriculum overview & course map | 3 modules x 6 lessons = 18 classes | Two tracks sharing one topic map: Pre-IGCSE (ages 10-12) & Pre-IB (ages 13-15)", 2)
rows = [
 ("Purpose", "A premium foundation course bridging beginners into IGCSE and IB Economics thinking and real-world application. Assumes no prior economics. Built as one conceptual spine delivered through two thinking tiers — Tier 1 / Pre-IGCSE (ages 10-12) discovers concepts through stories, games and simulations; Tier 2 / Pre-IB (ages 13-15) promotes the same topics into definitions, diagrams, calculations and evaluative writing — with a swappable exam adapter fitted on top (see the 3-layer architecture below)."),
 ("Design commitment 1", "Intuition before theory. Each concept starts from a story, game or real situation, then becomes a definition and a diagram. Across the tiers this is the promotion mechanic: the same topic, only the verbs upgraded — Tier 1 feels it, Tier 2 judges it."),
 ("Design commitment 2", "The exam adapter is drip-fed, never bolted on. Exam technique lives only in each lesson's Exam Adapter Slot (Tier 2), loaded lesson-by-lesson per the Exam Adapter Drip Map sheet: shared skills (definitions, diagrams, calculations) through M1, board-specific elements diverging from M2 L4, fully assembled by M3 L6. Sections 1-8 of every lesson stay exam-free."),
 ("Design commitment 3", "The nine IB key concepts as a spine: scarcity, choice, efficiency, equity, economic well-being, sustainability, change, interdependence, intervention. [Verified against the current IB DP Economics guide, first assessment 2022]"),
 ("Design commitment 4", "Dual alignment with honest flags. Every lesson is anchored subtopic-by-subtopic to Cambridge IGCSE 0455 (2027-2029) and the IB DP Economics Guide (2022) — see the two audit sheets. Core content is bulleted in priority order and boundary-marked: [ENRICH] = beyond one or both syllabuses, [IB SL CORE] = core for IB-bound only, [LOW] = minor, keep light, (!) = scope note. Nothing outside the current syllabuses is ever taught as core."),
 ("Design commitment 5", "One spine, single-point updates. Because both tiers hang off the same conceptual spine, a changing world requires exactly one edit: update the spine (a case, a news hook, a game) and it flows to both tiers automatically. This is what keeps the curriculum fluid in a fast-moving era without doubling maintenance."),
 ("3-layer architecture (design principle)", "This course does not split its tracks by exam board. It stacks a single conceptual spine, then two thinking tiers upon it — because the two boards share roughly 80% of their conceptual DNA, and a student at the point of entry does not yet know which board is their destination. Tier 1 'Draw the Map' (ages 10-12, the current Pre-IGCSE track): concepts discovered through games, simulations and stories, felt in the body before they are named. Tier 2 'Read the Map' (ages 13-15, the current Pre-IB track): the same topics, now carrying definitions, diagrams, calculations and evaluative writing. Layer 3, the exam adapter (not a separate module; fitted 1-2 years before the exam): the only layer where the boards genuinely diverge — an IGCSE adapter (3 AOs, no market-failure diagrams required) or an IB adapter (4 AOs, command terms, real-world-example essays) swaps on top of Tier 2. Each lesson's Exam-skill slot is the seed of this adapter layer. The public labels Pre-IGCSE / Pre-IB are retained, but internally these are thinking-stage tiers, not board tracks."),
 ("Two tracks (delivery)", "Tier 1 / Pre-IGCSE (ages 10-12): the same 18 topics, concrete and experiential — games, simulations, stories; simple 'trade-off curves' before formal PPCs; no heavy calculation. Tier 2 / Pre-IB (ages 13-15): full definitions, diagram discipline, AO-tagged objectives, calculations, command-term writing."),
 ("Modules", "1 - Microeconomics | 2 - Macroeconomics | 3 - Global Economics (6 lessons each). Independent, but recommended order is 1 -> 2 -> 3. v3 adds three lessons to close the audit gaps: M1 L5 Firms, Costs & Market Power; M2 L1 Money, Banking, Households & Workers; M3 L4 Inequality, Poverty & Population."),
 ("Each module", "Lessons 1-5 teach; Lesson 6 = review + mock examination + exam-skills clinic (Tier 2 / Pre-IB) or capstone simulation + showcase quiz (Tier 1 / Pre-IGCSE). The three Lesson 6s are the exam adapter's assembly checkpoints."),
 ("Framework basis", "IB side: current DP Economics guide (first assessment 2022; Tragakes 'Economics for the IB Diploma' structure): 4 syllabus units, 9 key concepts, 4 assessment objectives, Paper 1/2/3 formats. IGCSE side: Cambridge 0455 2027-2029 topics 1-6."),
]
r = 4
for k, v in rows:
    put(ws, r, [k, v], bold_first=True)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=GREY)
    r += 1
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 120

# ============================================================ Verification Notes
ws = wb.create_sheet("Verification Notes")
title_block(ws, "Verification against the current syllabuses (checked July 2026)",
            "[OK] = confirmed correct | [FIX] = corrected/updated in this version | [FLAG] = keep in mind when writing lesson materials", 4)
header_row(ws, 3, ["Status", "Item", "Finding", "Action taken in v2"])
vrows = [
 ("[OK]", "IB nine key concepts", "Scarcity, choice, efficiency, equity, economic well-being, sustainability, change, interdependence, intervention — matches the current IB DP Economics guide (first assessment 2022). No newer IB Economics guide has replaced it; Economics is not in the first-teaching-2027 update cycle.", "Kept as the curriculum spine, unchanged."),
 ("[OK]", "IB unit mapping", "Unit 1 Introduction, Unit 2 Microeconomics, Unit 3 Macroeconomics, Unit 4 The Global Economy. Module mapping (M1->Units 1-2, M2->Unit 3, M3->Unit 4) is correct.", "Unchanged."),
 ("[OK]", "IB assessment objectives & command terms", "AO1-AO4 and the command-term groupings in 'Exam Skills Progression' match the IB guide.", "Unchanged, but relabelled as IB AOs (see [FIX] below)."),
 ("[OK]", "IB paper formats", "Paper 1 extended response, Paper 2 data response, Paper 3 HL-only. Wording refined: Paper 3 is the HL 'policy paper' — calculations PLUS a policy recommendation, not calculations alone.", "Assessment Styles sheet wording updated."),
 ("[FIX]", "IGCSE topic labels", "The old 'Maps to IGCSE' labels ('the market system; individual as producer/consumer; the private firm; role of government') follow Edexcel/legacy naming. Cambridge 0455 (2027-2029) uses six topics: 1 The basic economic problem, 2 The allocation of resources, 3 Microeconomic decision makers, 4 Government and the macroeconomy, 5 Economic development, 6 International trade & globalisation.", "Structure sheet and module footers rewritten with the correct 2027-2029 topic names/numbers."),
 ("[FIX]", "IGCSE assessment objectives", "Cambridge 0455 (2027-2029) has only THREE AOs — AO1 Knowledge & understanding (43%), AO2 Analysis (47%), AO3 Evaluation (10%). The AO1-AO4 framework in this course map is the IB one; 'AO4 calculations' does not exist at IGCSE.", "Exam Skills Progression sheet now labels the AO framework as IB and adds the IGCSE AO note."),
 ("[FLAG]", "Market-failure diagrams at IGCSE", "The 2027-2029 syllabus states explicitly that demand & supply diagrams relating to market failure are NOT required at IGCSE. Externality/welfare-loss diagrams are IB-level.", "M1 L4 diagrams flagged: externality & welfare-loss diagrams = 'IB enrichment (not required for IGCSE 0455)'."),
 ("[FLAG]", "Content removed in 0455 2027-2029", "Removed or heavily reduced vs the 2026 syllabus: trade unions as a standalone topic; causes & forms of growth of firms; old 4.1 'role of government'; principles of taxation; changing patterns/level of employment.", "None of these were standalone lessons here — noted so future lesson materials don't over-teach removed content."),
 ("[OK]", "Everything else", "Lesson sequencing, key-concept map weighting, lesson template, question-style mapping: internally consistent and syllabus-appropriate.", "Preserved; columns added (Core Concepts, Essential Questions, two level tracks)."),
]
r = 4
for row in vrows:
    fills = [FIX if row[0] == "[FIX]" else (LIGHT_GOLD if row[0] == "[FLAG]" else None)]*4
    put(ws, r, list(row), fills=fills, bold_first=True)
    r += 1
for col, w in zip("ABCD", [7, 30, 70, 55]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A4"

# ============================================================ Structure
ws = wb.create_sheet("Structure")
title_block(ws, "Structure at a Glance", "How the three modules map to the IB units and the Cambridge IGCSE 0455 (2027-2029) topics — the foundational IGCSE concepts feeding forward into IB", 5)
header_row(ws, 3, ["Module", "Focus", "Lessons", "Maps to IB Unit", "Maps to IGCSE 0455 (2027-2029)"])
srows = [
 ("1 - Microeconomics", "Individual markets, prices, firms, government in markets", "6", "Units 1-2 (Introduction; Microeconomics)", "Topic 1 The basic economic problem; Topic 2 The allocation of resources; parts of Topic 3 Microeconomic decision makers"),
 ("2 - Macroeconomics", "Money, the whole economy, policy", "6", "Unit 3 (Macroeconomics)", "Topic 4 Government and the macroeconomy; parts of Topic 5 Economic development"),
 ("3 - Global Economics", "Trade, exchange, equity, development", "6", "Unit 4 (The Global Economy)", "Topic 6 International trade & globalisation; Topic 5 Economic development"),
]
r = 4
for row in srows:
    put(ws, r, list(row), bold_first=True); r += 1
for col, w in zip("ABCDE", [22, 34, 9, 34, 60]):
    ws.column_dimensions[col].width = w

# ============================================================ Key-Concept Map
ws = wb.create_sheet("Key-Concept Map")
title_block(ws, "Key-Concept Map (the IB spine)", "Where each of the nine key concepts is foregrounded. PP = primary lens for that module; P = revisited. Verified against the current IB guide.", 4)
header_row(ws, 3, ["Key concept", "Micro", "Macro", "Global"])
kc = [("Scarcity","PP","P","P"),("Choice","PP","P","P"),("Efficiency","PP","P","P"),
      ("Equity","P","PP","PP"),("Economic well-being","P","PP","P"),("Sustainability","P","P","PP"),
      ("Change","P","PP","P"),("Interdependence","P","P","PP"),("Intervention","PP","PP","P")]
r = 4
for row in kc:
    put(ws, r, list(row), bold_first=True)
    for i in (2,3,4):
        ws.cell(row=r, column=i).alignment = WRAP_C
    r += 1
for col, w in zip("ABCD", [24, 10, 10, 10]):
    ws.column_dimensions[col].width = w

# ============================================================ Module sheets
MOD_HEADERS = ["#", "Lesson title", "Core Concepts", "Essential Questions (2-3)", "Core content",
               "Key diagrams", "Exam-skill focus (Pre-IB)", "Pre-IGCSE track (ages 10-12)", "Pre-IB track (ages 13-15)"]
MOD_FILLS = [NAVY, NAVY, GOLD, GOLD, NAVY, NAVY, NAVY, GOLD, GOLD]
ROW_FILLS = [None, None, LIGHT_GOLD, LIGHT_GOLD, None, None, None, LIGHT_GOLD, LIGHT_GOLD]
WIDTHS = [4, 26, 30, 42, 48, 34, 30, 46, 46]

def module_sheet(name, title, subtitle, lessons, footer):
    ws = wb.create_sheet(name)
    title_block(ws, title, subtitle, 9)
    header_row(ws, 3, MOD_HEADERS, fills=MOD_FILLS)
    r = 4
    import re as _re
    for L in lessons:
        L = list(L)
        # blank separator line before the first flag block in Core content
        L[4] = _re.sub(r'\n(\[ENRICH|\[LOW\]|\[IB SL CORE|\(!\))', r'\n\n\1', L[4], count=1)
        put(ws, r, L, fills=ROW_FILLS)
        ws.cell(row=r, column=1).alignment = WRAP_C
        ws.cell(row=r, column=2).font = F(size=10, bold=True)
        ws.row_dimensions[r].height = 140 if 'Review' in L[1] else 275
        r += 1
    ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=9)
    c = ws.cell(row=r+1, column=1, value=footer)
    c.font = F(size=9, italic=True, color=NAVY); c.alignment = WRAP
    c.fill = PatternFill("solid", fgColor=CREAM)
    ws.row_dimensions[r+1].height = 40
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:I{r-1}"

m1 = [
 ("1", "The Economic Problem: Scarcity, Choice & Systems",
  "\u2022 Scarcity\n\u2022 Opportunity cost\n\u2022 Factors of production\n\u2022 Economic systems\nKC: Scarcity \u00b7 Choice \u00b7 Efficiency",
  "1. Why can't anyone \u2014 even the richest country \u2014 have everything it wants?\n2. What does every choice really cost?\n3. Who should decide what gets produced: markets, governments, or both?",
  "\u2022 Scarcity & the basic economic problem: finite resources vs infinite wants; needs vs wants; economic vs free goods\n\u2022 Opportunity cost: definition; examples across consumers, workers, firms and governments; its role in every decision\n\u2022 Factors of production (land, labour, capital, enterprise) and their rewards (rent, wages, interest, profit); causes of changes in their quantity & quality\n\u2022 PPC: drawing & interpretation; points under/on/beyond the curve; movements along (= opportunity cost); shifts (= growth)\n\u2022 The three resource-allocation questions: what / how / for whom\n\u2022 Economic systems \u2014 market vs planned vs mixed: definitions and arguments for & against (IGCSE 2.8 / 2.10)\n[ENRICH \u00b7 IB Unit 1] Economics as a social science; positive vs normative statements; methodology & pluralism\n[ENRICH \u00b7 IB] Circular flow of income \u2014 not in the IGCSE syllabus, but essential scaffolding for macro in M2\n[LOW] Brief history of economic thought \u2014 one story, no depth",
  "Production Possibilities Curve (PPC); circular flow model",
  "AO1 command terms; how to define precisely; reading a PPC",
  "Desert-island survival game: rank what to salvage, feel scarcity. 'Give up to get' trading round makes opportunity cost physical. Draw a simple 2-good 'trade-off curve' (pizza vs robots) instead of a formal PPC. Story of three villages (market / planned / mixed) \u2014 students vote on which they'd live in and defend it. 'Fact or opinion?' card sort seeds positive vs normative.",
  "Precise AO1 definitions; full PPC analysis (points on/inside/outside curve, shifts = growth); 2-sector circular flow; compare real economies (e.g. US / North Korea / Singapore) as systems; positive vs normative statements; first structured 'define + explain' written answers."),
 ("2", "Demand, Supply & Market Equilibrium",
  "\u2022 Demand & supply\n\u2022 The price mechanism\n\u2022 Equilibrium & disequilibrium\nKC: Efficiency \u00b7 Interdependence \u00b7 Change",
  "1. How does a price 'know' how to change without anyone commanding it?\n2. What is a price actually telling buyers and sellers?\n3. What happens when a market is out of balance \u2014 and who fixes it?",
  "\u2022 Law of demand & law of supply; individual vs market demand/supply; drawing & interpreting both diagrams\n\u2022 Shifts OF a curve vs movements ALONG it \u2014 the single most common exam error; the full determinant taxonomy for each curve\n\u2022 Market equilibrium & disequilibrium using BOTH schedules (tables) and curves; shortages & surpluses\n\u2022 The price mechanism: how prices signal, ration and incentivise \u2014 answering what / how / for whom\n\u2022 Causes & consequences of price changes; the effect on sales (feeds the revenue link in L3)\n[ENRICH \u00b7 IB SL] Consumer & producer surplus \u2014 not required at IGCSE\n[ENRICH \u00b7 IB HL 2.4] Behavioural-economics seed: biases, nudges, the critique of rational maximisation",
  "D/S diagram; shifts vs movements along; equilibrium adjustment",
  "Drawing & fully labelling D/S; distinguishing shift from movement",
  "Live pit-market simulation: half the class buyers, half sellers with secret price cards \u2014 watch the price converge on its own. Sneaker-drop / concert-ticket stories for shortage. Plot the class's own trading data as the first 'demand curve'. No shifter taxonomy \u2014 just 'what would make you want more of this?' Bonus: a 'trick your brain' nudge experiment (behavioural econ seed).",
  "Full D/S diagram discipline (labels, arrows, equilibria e1->e2); complete determinant taxonomy; schedules AND curves; shift-vs-movement drills; consumer/producer surplus shading (IB enrichment); behavioural-econ enrichment discussion; write cause->effect chains using 'explain' and 'analyse'."),
 ("3", "Elasticity",
  "\u2022 Responsiveness (elasticity)\n\u2022 PED & PES\n\u2022 The revenue link\nKC: Change \u00b7 Choice \u00b7 Efficiency",
  "1. Why do petrol prices swing wildly while haircut prices barely move?\n2. When does raising your price actually lose you money?\n3. How can a government use elasticity to change what we buy?",
  "\u2022 PED: definition, formula, calculation with full working; the five-value vocabulary (perfectly inelastic -> inelastic -> unitary -> elastic -> perfectly elastic) with a diagram for each\n\u2022 PED and revenue \u2014 the most-tested application: effect of price changes on consumer expenditure & firms' revenue, shown BOTH on a diagram and as a calculation\n\u2022 Determinants of PED; significance for consumer, firm and government decisions\n\u2022 PES: definition, formula, calculation, five-value interpretation, determinants\n[ENRICH \u00b7 IB] YED: normal vs inferior goods; interpreting income elasticity\n[ENRICH \u00b7 IB/HL] Tax-incidence intro: who really pays a tax depends on elasticity\n[LOW] XED \u2014 in NEITHER current syllabus (IB removed it in the 2022 guide; IGCSE never had it). Mention only if a student asks",
  "Elastic vs inelastic curves; revenue effects",
  "Calculations (show your working); interpreting values \u2014 maps to IGCSE AO2 / IB AO4",
  "'Stretchy vs stiff' framing with a rubber band. Lemonade-stand pricing game: teams pick prices for water-in-a-desert vs one-of-many candy brands and discover revenue effects by playing. Sort real products onto an 'easy to give up <-> impossible to give up' line. Only PED intuition \u2014 no formulas.",
  "Full PED/PES/YED with calculations and sign interpretation; determinants; total-revenue rule proven numerically; tax-incidence intro (IB/HL enrichment); calculation technique \u2014 show working, interpret the value, state the implication for a firm or government."),
 ("4", "Government Intervention & Market Failure",
  "\u2022 Market failure\n\u2022 Externalities & public goods\n\u2022 Intervention & its limits\nKC: Intervention \u00b7 Efficiency \u00b7 Equity",
  "1. When do markets fail \u2014 and how would we know?\n2. Who pays for pollution if the polluter doesn't?\n3. Does government intervention fix markets, or sometimes break them further?",
  "\u2022 Market failure: definition + the full term set \u2014 public goods, merit & demerit goods, private / external / social costs & benefits, monopoly (newly listed as a market-failure term at IGCSE 2027-29)\n\u2022 Causes & consequences of misallocation: over-consumption of demerit / external-cost goods; under-consumption of merit / external-benefit goods; non-provision of public goods; monopoly restricting supply and raising price\n\u2022 Intervention on D/S diagrams (drawing REQUIRED at IGCSE): maximum & minimum prices, indirect taxes, subsidies \u2014 each with advantages & disadvantages\n\u2022 The wider toolkit (definitions + pros/cons, no diagrams): regulation, privatisation & nationalisation, direct provision, quotas\n\u2022 Market vs mixed economic system: arguments for and against each\n[ENRICH \u00b7 IB] Externality & welfare-loss diagrams \u2014 NOT required at IGCSE 2027-29\n[ENRICH \u00b7 IB] Common pool resources & the tragedy of the commons (the sustainability link)\n[LOW] Black-market consequences of price controls \u2014 one example, no depth",
  "Price controls on D/S; tax & subsidy on D/S. Externality & welfare-loss diagrams = IB enrichment (NOT required for IGCSE 0455 2027-2029)",
  "Intro to evaluation: discuss / evaluate; stakeholder analysis (IGCSE AO3 = evaluation, 10% weighting)",
  "Factory-by-the-river role play: factory owner, fishermen, townspeople negotiate \u2014 externalities felt before named. Fishing-lake game: shared bowl of goldfish crackers depletes in rounds \u2014 tragedy of the commons discovered, not lectured. 'Free-rider' classroom experiment for public goods. Debate: 'Should candy be taxed like cigarettes?'",
  "Formal intervention toolkit on D/S diagrams (ceiling, floor, tax, subsidy) plus regulation/privatisation/nationalisation/direct provision/quotas with stakeholder tables (winners/losers); externality & CPR classification; merit/demerit/public goods and monopoly abuse with real policy cases; first full 'discuss' paragraph with a supported judgement."),
 ("5", "Firms, Costs & Market Power  [NEW in v3]",
  "\u2022 Costs, revenue & profit\n\u2022 Economies of scale\n\u2022 Competition vs monopoly\nKC: Efficiency \u00b7 Choice \u00b7 Interdependence",
  "1. Why do some firms want to become huge \u2014 and when does big become a problem?\n2. Where does profit actually come from?\n3. Is one seller always worse for us than many?",
  "\u2022 Costs of production: TC, ATC, FC, AFC, VC, AVC \u2014 definitions, calculations, and diagrams showing how costs change with output\n\u2022 Revenue: TR & AR \u2014 definitions, calculations, and how sales volume drives revenue\n\u2022 Economies & diseconomies of scale (internal AND external) with the ATC diagram \u2014 newly explicit at IGCSE 2027-29\n\u2022 Objectives of firms: survival, social welfare, profit maximisation, growth\n\u2022 Types of firms: sector (primary/secondary/tertiary), ownership (private/public), advantages & disadvantages of small vs large\n\u2022 Mergers: horizontal, vertical, conglomerate \u2014 definitions, real examples, pros & cons\n\u2022 Competitive vs monopoly markets: characteristics and effects on price, quality, choice, profit (NO diagrams or formal theory at IGCSE)\n\u2022 Firms & production: demand for factors of production (product demand, factor prices, availability, productivity); labour- vs capital-intensive production; production vs productivity; effect of investment on productivity\n[ENRICH \u00b7 IB HL 2.11] Market-structure spectrum preview (perfect competition -> monopoly)\n[ENRICH \u00b7 IB HL 2.10] Asymmetric information: the 'lemons' problem",
  "Average total cost (ATC) diagram for economies/diseconomies of scale (newly explicit in IGCSE 2027-29); cost-curve sketches. IB HL enrichment: market-structure spectrum",
  "Cost & revenue calculations (show working); advantages/disadvantages tables; applying 'analyse' to a firm scenario",
  "Cookie-factory simulation: teams 'produce' in rounds, hire more bakers and buy machines \u2014 feel average cost fall (economies of scale) then chaos as the factory floor jams (diseconomies). Merger negotiation round: two rival factories bargain \u2014 horizontal or vertical? 'Only lemonade stand in town' role play: the monopolist sets prices while the town reacts. Used-bikes 'lemons' mini-game as a teaser.",
  "Full cost/revenue vocabulary with calculations (TC/ATC/FC/AFC/VC/AVC, TR/AR); ATC diagram drawn to exam standard; merger types with real cases (e.g. Disney-Pixar vs a supermarket buying a farm); competitive-vs-monopoly comparison tables; structured 'analyse the effect on price, quality, choice and profit' answers; HL teaser on market structures."),
 ("6", "Review + Mock Examination + Exam Skills",
  "\u2022 Micro synthesis\n\u2022 Exam craft\nKC: all micro-primary concepts revisited",
  "1. How do scarcity, prices, elasticity, firms and intervention fit into one story of a market?\n2. What does an examiner actually reward \u2014 and what wastes your time?",
  "\u2022 Full micro synthesis: scarcity -> prices -> elasticity -> firms -> intervention as one connected story\n\u2022 Concept-connection mapping & misconception clinic (shift vs movement; 'price rises so demand falls' wording; cost vs price)\n\u2022 Mock examination: MCQ + short answers + data response + extended response\n\u2022 Mark-scheme literacy and timing strategy",
  "All of the above",
  "Timing, planning, mark-scheme literacy",
  "Capstone: 'Market Day' simulation \u2014 students run stalls, a 'government' levies taxes and sets price rules, one stall wins a 'monopoly licence'; debrief maps every event to a concept. Team quiz-show review (buzzers, concept-connection bonus round). Short, friendly written check \u2014 no exam pressure.",
  "Full mock under timed conditions (MCQ + short + data response + extended); self-marking against a mark scheme; misconception clinic; exam-strategy debrief."),
]
m1_footer = ("IGCSE 0455 (2027-2029) alignment: Topic 1 The basic economic problem; Topic 2 The allocation of resources (market-failure D/S diagrams not required at IGCSE); Topic 3.4-3.7 Firms, production, costs & types of markets. "
             "IB preview: Units 1-2 incl. behavioural economics seed; HL-flagged enrichment (consumer/producer surplus, welfare-loss diagrams, tax incidence, market power, asymmetric information).")

m2 = [
 ("1", "Money, Banking, Households & Workers  [NEW in v3]",
  "\u2022 Money & banking\n\u2022 Household finance\n\u2022 Wage determination\nKC: Interdependence \u00b7 Choice \u00b7 Equity",
  "1. Why does everyone accept a piece of paper \u2014 or a number on a screen \u2014 as valuable?\n2. What does a bank actually do with your money?\n3. Why does a footballer earn more than a nurse?",
  "\u2022 Money: forms, functions, characteristics \u2014 why money must hold its promises\n\u2022 Banking: role & importance of the central bank vs commercial banks (the lever that monetary policy pulls in L5)\n\u2022 Wage determination: demand & supply of labour, relative bargaining power, government policy \u2014 with the labour-market AND minimum-wage diagrams (drawing newly explicit at IGCSE 2027-29)\n\u2022 Reasons for wage differences: skill level, sector, discrimination, public vs private, bargaining strength\n\u2022 Households: influences on spending, saving & borrowing \u2014 income, interest rate, confidence, age, culture\n\u2022 Choice of occupation: wage vs non-wage factors\n\u2022 Division of labour / specialisation: definition, advantages & disadvantages\n\u2022 Mobility of labour (occupational & geographical): causes and consequences of changes\n[LOW] Trade unions \u2014 no longer a standalone IGCSE topic (2027-29); cover only as one bargaining-power influence inside wage determination",
  "Labour-market D/S diagram; national minimum wage diagram (newly explicit in IGCSE 2027-29)",
  "Transferring D/S reasoning to a new market (labour); diagram transfer; 'explain the difference in wages' structured answers",
  "Invent-your-own-currency game: teams design a money (shells? stickers? bottle caps?) and stress-test it through trades \u2014 the six characteristics of good money get DISCOVERED. Mini-bank simulation: deposits, loans and a (gentle) bank run. Salary auction: class bids for jobs described only by conditions \u2014 why wages differ emerges from the bidding. Family budget role-play cards (age & culture shape choices).",
  "Money functions & characteristics precisely; central vs commercial bank roles; savings/borrowing analysis with interest-rate reasoning; labour-market D/S diagrams incl. NMW to exam standard; structured multi-cause analysis of wage differences; bridge forward: 'the interest rate lever returns in Lesson 5 as monetary policy'."),
 ("2", "Measuring the Economy",
  "\u2022 GDP & national income\n\u2022 Real vs nominal\n\u2022 Living standards & their limits\nKC: Economic well-being \u00b7 Change",
  "1. Can one number capture how well a whole country is doing?\n2. If GDP rises, is everyone actually better off?\n3. What should we measure if what we care about is well-being?",
  "\u2022 GDP: definition; nominal vs real; GDP per capita; growth-rate calculations with full working\n\u2022 Limitations of GDP as a living-standards measure (distribution, non-market activity, environment) \u2014 with HDI as the counterpart indicator\n\u2022 The economic (business) cycle: the stages and what moves an economy through them\n\u2022 GNI vs GDP: the distinction and when each is used\n[ENRICH \u00b7 IB] Circular flow with injections & leakages; GDP-deflator intuition\n[LOW] National-accounting technicalities beyond the above \u2014 do not go deeper",
  "Business cycle; circular flow (with injections/leakages \u2014 IB framing)",
  "Calculations: real GDP, growth rate, per capita",
  "'Country in a box' game: each team is a mini-nation producing paper goods; count 'GDP' each round, then reveal what GDP missed (a team that polluted, a team where one member got everything). Happiness vs income country-card matching. Business cycle as a rollercoaster storyboard.",
  "GDP vs GNI precisely; real vs nominal with deflator intuition; per-capita and growth-rate calculations with working; circular flow with injections/leakages; structured critique of GDP (distribution, non-market activity, environment) + alternative indicators (HDI)."),
 ("3", "The AD-AS Model",
  "\u2022 Aggregate demand & supply\n\u2022 Macro equilibrium\n\u2022 Schools of thought\nKC: Interdependence \u00b7 Change",
  "1. What moves an entire economy up or down?\n2. Why do economists disagree about what the supply side looks like (Keynesian vs monetarist)?",
  "\u2022 Components of aggregate demand \u2014 C + I + G + (X-M) \u2014 and what shifts each\n\u2022 Macroeconomic equilibrium and diagram discipline (Price Level / Real Output \u2014 never P/Q)\n\u2022 Short-run aggregate supply and its shifts\n\u2022 Output gaps: identifying negative & positive gaps on the diagram\n\u2022 Keynesian vs monetarist/new-classical AS shapes \u2014 and why economists disagree\n(!) SCOPE NOTE: the AD-AS model itself is IB syllabus content. IGCSE 0455 never requires AD-AS diagrams \u2014 for IGCSE-bound students this lesson is conceptual scaffolding (the 'total demand / total supply' language the IGCSE syllabus does use), which pays off in L4-L5",
  "AD-AS diagram (both AS shapes); shifts of AD and AS. NOTE: the AD-AS diagram is IB-level \u2014 not required anywhere in IGCSE 0455",
  "Correct AD-AS labelling ('Price Level' / 'Real Output', never P/Q)",
  "Build the economy as a class machine: four groups ARE C, I, G and (X-M); events ('a scary news week', 'government builds a stadium') make groups step forward/back and the class reads total demand. Keynes vs Hayek intro via the rap-battle video + 'whose side are you on?' discussion. No formal AS shapes.",
  "Full AD-AS model with both AS shapes and why they differ; shift analysis for each AD component; output gaps; labelling discipline drilled (Price Level / Real Output); chain-of-reasoning paragraphs from a shock to the new equilibrium."),
 ("4", "Growth, Unemployment & Inflation",
  "\u2022 Inflation & CPI\n\u2022 Unemployment types\n\u2022 Growth & recession\nKC: Change \u00b7 Economic well-being \u00b7 Sustainability",
  "1. Can an economy grow forever?\n2. Which is worse \u2014 unemployment or inflation \u2014 and for whom?\n3. Why is deflation (falling prices) scarier than it sounds?",
  "\u2022 Inflation: definition; CPI measurement & the inflation-rate calculation; demand-pull vs cost-push; consequences (savers, lenders, borrowers; consumers, workers, firms); policies to control it\n\u2022 Unemployment: definitions (incl. full employment); labour-force-survey measurement and the rate formula; the four types (frictional, structural, cyclical, seasonal \u2014 seasonal newly explicit at IGCSE); consequences for individual / firms / government / economy; policies to reduce it\n\u2022 Economic growth: definition; measurement by real GDP; causes & consequences (advantages AND disadvantages); policies to promote it\n\u2022 Recession: definition, causes and consequences for each stakeholder \u2014 newly explicit at IGCSE 2027-29\n[ENRICH \u00b7 IB] Deflation & the deflationary spiral (removed from IGCSE 2027-29)\n[ENRICH \u00b7 IB] Growth framed on the PPC / LRAS (IGCSE dropped PPC references here)\n[ENRICH \u00b7 IB HL] The short-run Phillips curve trade-off",
  "Growth on PPC & LRAS (IB framing); deflation via AD/AS (IB); Phillips curve (IB HL enrichment)",
  "Calculations: unemployment rate, inflation rate from CPI",
  "Inflation felt through a 'shrinking pocket-money' shopping game across three rounds of rising price tags. Unemployment types as story cards (the robot took my job / factory moved / just graduated / ski instructor in summer) sorted by students. Hyperinflation photo stories (wheelbarrows of cash) \u2014 why money must keep its promise.",
  "Types of unemployment with policy implications; CPI construction and inflation-rate calculation; demand-pull vs cost-push on AD-AS; recession causes/consequences; deflation spiral analysis (IB); growth as PPC/LRAS shift; HL enrichment: short-run Phillips curve trade-off."),
 ("5", "Demand-side & Supply-side Policy",
  "\u2022 Macro aims & conflicts\n\u2022 Fiscal & monetary policy\n\u2022 Supply-side policy\nKC: Intervention \u00b7 Equity \u00b7 Sustainability",
  "1. Who really steers the economy \u2014 the government, the central bank, or nobody?\n2. Why does fixing one macro problem so often worsen another?\n3. What would make growth 'sustainable'?",
  "\u2022 The six macroeconomic aims BY NAME (growth, full employment, stable prices, balance-of-payments stability, redistribution of income, environmental sustainability) and the three classic conflicts between them\n\u2022 Fiscal policy: the government budget \u2014 deficit/surplus definitions AND calculations; main areas & reasons for spending; reasons for taxation; tax classifications (progressive / regressive / proportional; direct / indirect); impact on each stakeholder; fiscal measures and their effect on each aim\n\u2022 Monetary policy: definitions of money supply & monetary policy; the three measures \u2014 interest rate, money supply, foreign-exchange rate; how each serves the macro aims\n\u2022 Supply-side policy: the full measures list (education & training, infrastructure spending, labour-market reform, lower direct taxes, deregulation, incentives to work & invest, privatisation) and effects on the aims\n\u2022 Policy matching: which policy for which problem \u2014 closing the 'policies to promote growth / reduce unemployment / control inflation' threads from L4\n[ENRICH \u00b7 IB HL] Sustainable level of government debt; the Keynesian multiplier\n[LOW] Exchange-rate-channel mechanics of monetary policy \u2014 name it, don't model it",
  "AD/AS shifts from each policy (IB framing; IGCSE argues the same chains verbally)",
  "Full evaluation of a policy (strengths / limitations / judgement)",
  "'You are the Central Bank' game: teams set interest rates each round reacting to news cards; economy dial (spinner) responds; leaderboard for stability, not growth. Budget game: allocate 100 coins of government spending AND choose who pays which taxes \u2014 fairness argument erupts naturally (progressive vs flat felt, not defined). Trade-off felt: every fix card creates a new problem card.",
  "Fiscal vs monetary mechanisms through AD; budget deficit/surplus calculations; tax classifications with incidence-fairness discussion; transmission of interest rates; supply-side policies and LRAS; conflicts (growth vs inflation vs environment) argued both ways; full evaluation paragraphs: policy -> strengths -> limitations -> context-dependent judgement."),
 ("6", "Review + Mock Examination + Exam Skills",
  "\u2022 Macro synthesis\n\u2022 Data-response craft\nKC: all macro-primary concepts revisited",
  "1. How do the macro pieces connect into one working model of an economy?\n2. How do you turn a table of data into an argument?",
  "\u2022 Full macro synthesis: money -> measurement -> model -> problems -> policy\n\u2022 Concept connections & misconception clinic (real vs nominal; deficit vs debt; money vs income)\n\u2022 Mock with an IB Paper 2-style data response / IGCSE Section A format\n\u2022 Technique: extracting and quoting data as evidence",
  "All of the above",
  "Data-response technique; extracting evidence from a source",
  "Capstone: 'Save the Economy' team simulation \u2014 a scripted crisis unfolds over 5 news bulletins; teams choose policies (now including the banking lever from L1) and see consequences; debrief maps decisions to concepts. Quiz-show review with a graph-drawing lightning round.",
  "Full macro mock featuring an IB Paper 2-style data response; technique for quoting/using data as evidence; misconception clinic; timing and answer-planning drills."),
]
m2_footer = ("IGCSE 0455 (2027-2029) alignment: Topic 3.1-3.3 Money & banking, households, workers; Topic 4 Government and the macroeconomy (aims, fiscal/monetary/supply-side policy, growth, unemployment, inflation); parts of Topic 5. "
             "IB preview: Unit 3 Macroeconomics; HL enrichment (Phillips curve, multiplier intuition, sustainable government debt).")

m3 = [
 ("1", "International Trade & Comparative Advantage",
  "\u2022 Specialisation\n\u2022 Free trade\n\u2022 Gains from trade\nKC: Interdependence \u00b7 Efficiency",
  "1. Why do countries trade instead of making everything themselves?\n2. Can trade really make BOTH sides richer at the same time?",
  "\u2022 Specialisation by country: definition; the basis for it (best resource allocation / low-cost production); advantages & disadvantages\n\u2022 Free trade: definition; advantages & disadvantages\n\u2022 Gains from trade demonstrated numerically \u2014 both sides can win\n[ENRICH \u00b7 IB HL] The formal absolute vs comparative advantage model with opportunity-cost tables\n[LOW] Terms-of-trade intuition \u2014 one sentence, not a topic",
  "Gains-from-trade demonstration; comparative-advantage numerical example (IB HL)",
  "Structured explanation; numerical reasoning",
  "Trading-card economy: each team can 'produce' only certain shapes/colours cheaply; free-trade rounds vs no-trade rounds \u2014 score totals prove gains from trade. 'Why doesn't Messi mow his own lawn?' as comparative advantage. Map activity: where everything in your backpack was made.",
  "Absolute vs comparative advantage with opportunity-cost tables (HL enrichment); numerical worked examples both directions; terms-of-trade intuition; structured multi-step explanation writing; free-trade arguments with real bloc examples."),
 ("2", "Protectionism",
  "\u2022 Tools of protection\n\u2022 Reasons for restriction\n\u2022 Winners & losers\nKC: Intervention \u00b7 Interdependence \u00b7 Equity",
  "1. Who wins and who loses when a country blocks trade?\n2. Is protecting local jobs worth higher prices for everyone else?",
  "\u2022 The tools of protection: tariffs, import quotas, subsidies, embargoes (newly explicit at IGCSE 2027-29), administrative barriers\n\u2022 Reasons for restrictions \u2014 the full IGCSE list: infant (sunrise) industries; declining (sunset) industries; strategic industries; anti-dumping; reducing a current-account deficit; raising tax revenue; restricting demerit-good imports; environmental sustainability\n\u2022 Consequences: impact on the home country AND its trading partners; advantages & disadvantages of restricting free trade\n\u2022 Each tool analysed on a D/S diagram\n[ENRICH \u00b7 IB] The tariff diagram with welfare areas\n[LOW] Administrative-barrier case detail \u2014 one example suffices",
  "Each protectionist tool on D/S; tariff diagram with welfare areas (IB-level detail); subsidy to domestic producers",
  "Weighing arguments for/against; 'examine' / 'discuss' technique",
  "Continue the trading-card economy: one nation imposes a tariff, another a full embargo \u2014 feel the ripple. Courtroom drama: 'The People vs The Tariff' with student lawyers for consumers, workers, firms, government. Real story hooks: chip wars, K-food export bans, sneaker tariffs.",
  "Each protectionist tool on a D/S diagram; tariff diagram with welfare areas (IB enrichment); the full IGCSE reasons-for-restrictions list assessed case by case; balanced 'discuss' essays with explicit winners/losers and a judgement."),
 ("3", "Exchange Rates & the Balance of Payments",
  "\u2022 Currency markets\n\u2022 Appreciation & depreciation\n\u2022 The current account\nKC: Interdependence \u00b7 Change",
  "1. What makes a currency rise or fall \u2014 and who celebrates each move?\n2. Is a weak currency bad news? For whom?\n3. What does a current-account deficit really tell us?",
  "\u2022 Floating exchange-rate determination: demand & supply of a currency; the equilibrium rate; definitions of floating rate, appreciation, depreciation\n\u2022 Causes of fluctuations \u2014 changes in export/import demand, interest rates, speculation \u2014 plus the full list of reasons currencies are bought & sold: trade, speculation, government intervention, profit/interest/dividend flows, workers' remittances, capital investment\n\u2022 Consequences of exchange-rate changes for export & import prices and demand\n\u2022 Current account of the balance of payments: the four components (trade in goods, trade in services, primary income, secondary income); deficit/surplus calculations; causes; consequences for GDP, employment, inflation and the exchange rate; policies for stability\n\u2022 Cross-currency price conversions with working\n[ENRICH \u00b7 IB HL] Fixed & managed exchange-rate systems (removed from IGCSE 2027-29)\n[ENRICH \u00b7 IB] The full balance of payments incl. the financial account",
  "Currency market (demand/supply of a currency); the four exchange-rate change diagrams",
  "Calculations: converting prices across currencies; current-account deficit/surplus calculations",
  "Currency market as a live game: two-team foreign-exchange booth trading 'won' for 'dollars' as news cards land (K-pop world tour! / tourists stay home / grandma sends remittances). Travel-budget challenge: same trip priced in 3 currencies across two dates. Big Mac Index treasure hunt.",
  "Currency D/S diagrams for all four movement cases; the full reasons-for-buying/selling list; effects on households/exporters/importers; current-account structure with calculations; IB extension: financial account & the whole BoP; cross-currency price calculations with working."),
 ("4", "Inequality, Poverty & Population  [NEW in v3]",
  "\u2022 Poverty (absolute & relative)\n\u2022 Redistribution\n\u2022 Population dynamics\nKC: Equity (primary anchor) \u00b7 Economic well-being \u00b7 Change",
  "1. How unequal is too unequal \u2014 and how would we even measure it?\n2. Can a rich country still have poverty?\n3. Is a growing population a burden or an engine?",
  "\u2022 Poverty: absolute vs relative \u2014 definitions and the difference between them; causes (unemployment, low wages, illness, age, environmental factors)\n\u2022 Policies to alleviate poverty & redistribute income: promoting growth, improved education, improved healthcare provision, more generous state benefits, progressive taxation, the NMW\n\u2022 Living standards & income distribution: reasons for differences within AND between countries\n\u2022 Population: definitions of birth rate, death rate, net migration / immigration / emigration; why each varies between countries; the concept of an optimum population; effects of changes in population size and age/gender structure\n[IB SL CORE \u00b7 beyond IGCSE] Lorenz curve & Gini coefficient \u2014 enrichment for IGCSE-bound students, core for IB-bound (IB 3.4)\n[ENRICH \u00b7 IB HL 2.12] The market's inability to achieve equity\n[LOW] Gender-distribution effects \u2014 acknowledge, don't expand",
  "Lorenz curve & Gini (IB; not required at IGCSE); poverty-cycle flowchart",
  "Interpreting distribution data; Gini interpretation; evaluating redistribution policies",
  "Unequal-start candy economy: 10 sweets shared 6-2-1-1-0 \u2014 the class plots its own 'sharing curve' (a Lorenz curve, discovered). Poverty-trap board game: some players start with no dice. 'Village of 100 people' population simulation: age cards, births, migrations \u2014 the village suddenly grows old and the pension debate begins.",
  "Lorenz curve & Gini coefficient formally, with country comparisons; absolute vs relative poverty precisely; cause-by-cause poverty analysis; full evaluation of redistribution policies (incl. the progressive-tax link back to M2 L5); population-structure analysis with real demographic data (Korea & Japan ageing as live cases)."),
 ("5", "Globalisation, Integration & Development",
  "\u2022 Growth vs development\n\u2022 Globalisation & MNCs\n\u2022 Sustainability\nKC: Sustainability \u00b7 Equity \u00b7 Economic well-being",
  "1. Why are some countries rich and others poor?\n2. Is growth the same thing as development?\n3. Can the whole world develop without wrecking the planet?",
  "\u2022 Growth vs development: the distinction; indicators \u2014 real GDP per head and HDI (components, advantages & disadvantages of each)\n\u2022 Causes & consequences of international development differences: income, productivity, population growth, sector balance (primary/secondary/tertiary), saving & investment, education, healthcare, natural resources\n\u2022 Globalisation: definition; causes of change (trade restrictions, transport costs, communication costs, MNC movement); effects on trade, competition, the environment, migration, income distribution, development\n\u2022 Multinational companies: advantages & disadvantages to host AND home countries (explicit at IGCSE 2027-29)\n\u2022 Sustainability & climate economics \u2014 a theme now integrated throughout the 2027-29 syllabus\n[ENRICH \u00b7 IB] Trade blocs & the integration ladder (preferential -> FTA -> customs union -> common market -> monetary union; depth = HL)\n[ENRICH \u00b7 IB] The roles of the WTO / IMF / World Bank (not named in the IGCSE syllabus)\n[ENRICH \u00b7 IB 4.10] Development strategies by name: trade strategies, diversification, market-based vs interventionist, foreign aid, FDI, institutional change",
  "Development indicators; barriers-to-development flowchart",
  "Real-world examples as evidence; justify / recommend",
  "'Two babies born today' parallel-lives story (born in Norway vs a low-income country) \u2014 students discover what development means before defining it. MNC role play: a giant company knocks on a small country's door \u2014 town hall debate. Country-detective: identify nations from indicator cards.",
  "Growth vs development rigorously distinguished; HDI and multidimensional indicators; MNC host/home analysis; integration ladder named; barriers to development (poverty cycle, debt, institutions); development strategies assessed; roles of WTO/IMF/World Bank critically assessed; 'justify/recommend' answers deploying real-world evidence."),
 ("6", "Review + Mock Examination + Exam Skills",
  "\u2022 Global synthesis\n\u2022 Essay craft & RWE\nKC: all global-primary concepts revisited",
  "1. How do trade, currencies, inequality and development connect into one global story?\n2. How do you make a real-world example win marks instead of decorating a paragraph?",
  "\u2022 Full global synthesis: trade -> money across borders -> equity -> shared development\n\u2022 Concept connections & misconception clinic\n\u2022 Mock building to two full IB Paper 1-style (a)+(b) essays / IGCSE Section B extended responses\n\u2022 Deploying real-world examples as evidence \u2014 the final skill check",
  "All of the above",
  "Full essay structure; deploying a real-world example in part (b)",
  "Capstone: 'World Summit' simulation \u2014 each team represents a country (with unequal starting hands carried over from L4) negotiating a trade & climate deal; success scored on prosperity AND equity AND sustainability. Course-wide concept-connection wall: students physically link 18 lessons with string.",
  "Two full IB Paper 1-style (a)+(b) essays with planning time; real-world-example bank building; essay-structure mastery (claim -> analysis -> evaluation); final exam-strategy synthesis across all three papers' styles."),
]
m3_footer = ("IGCSE 0455 (2027-2029) alignment: Topic 6 International trade & globalisation; Topic 5 Economic development incl. 5.2 Poverty & 5.3 Population. "
             "IB preview: Unit 4 The Global Economy + Unit 3.4 Inequality & Poverty (SL core — Lorenz/Gini); Equity key concept now content-anchored.")

module_sheet("M1 Microeconomics", "Module 1 - Microeconomics (6 Lessons)",
             "Individual markets, prices, firms and government | Core content is BULLETED IN PRIORITY ORDER (top = most important). [ENRICH] = IB/HL enrichment beyond one or both syllabuses. [LOW] = minor \u2014 keep light, do not over-invest (retained for completeness). (!) = scope note.", m1, m1_footer)
module_sheet("M2 Macroeconomics", "Module 2 - Macroeconomics (6 Lessons)",
             "Money, the whole economy and policy | Core content is BULLETED IN PRIORITY ORDER (top = most important). [ENRICH] = IB/HL enrichment beyond one or both syllabuses. [LOW] = minor \u2014 keep light, do not over-invest. (!) = scope note.", m2, m2_footer)
module_sheet("M3 Global Economics", "Module 3 - Global Economics (6 Lessons)",
             "Trade, exchange, equity and development | Core content is BULLETED IN PRIORITY ORDER (top = most important). [ENRICH] = IB/HL enrichment. [IB SL CORE] = core for IB-bound, enrichment for IGCSE-bound. [LOW] = minor \u2014 keep light. ", m3, m3_footer)

# ============================================================ Lesson Template (3-layer aligned)
ws = wb.create_sheet("Lesson Template")
title_block(ws, "Standard Lesson Template — aligned to the 3-Layer Architecture",
    "Every one of the 18 lessons carries all 12 sections. Column 3 shows which architectural layer each section belongs to: SPINE = shared conceptual core (write once, flows to both tiers) | TIER 1 / TIER 2 = delivery layer (same topic, different verbs) | ADAPTER = the exam layer, drip-fed per lesson via the Exam Adapter Drip Map sheet. When updating content, edit SPINE sections once; when differentiating, touch only the tier columns; when changing exam technique, touch only the ADAPTER slot.", 5)

LAYER_FILLS = {"SPINE": CREAM, "TIER SPLIT": "DCE9F7", "ADAPTER": LIGHT_GOLD}
LAYER_FONTS = {"SPINE": NAVY, "TIER SPLIT": NAVY, "ADAPTER": NAVY}

header_row(ws, 3,
    ["#", "Section (present in every lesson)", "Layer", "Tier 1 (10-12) delivery — Draw the Map", "Tier 2 (13-15) delivery — Read the Map"],
    fills=[NAVY, NAVY, GOLD, "4A7AB5", "5E8C61"])

trows = [
 ("1", "Lesson Title", "SPINE",
  "Same title, both tiers.",
  "Same title, both tiers."),
 ("2", "Learning Objectives (4-8, measurable)", "TIER SPLIT",
  "Written as 'I can…' statements — no AO tags.",
  "Each objective tagged with IB AO1-AO4 (the AO tags are what the adapter layer will later grab onto)."),
 ("3", "Key Concepts — with the IB key-concept links", "SPINE",
  "The concept named in one child-friendly sentence, linking the lesson to a key concept every time.",
  "Full key-concept framing: which of the nine lenses this lesson feeds, and how it connects to prior lessons."),
 ("4", "Detailed Lesson Content — full teaching notes: stories, analogies, discussion prompts, everyday examples (prose, not bullet dumps)", "SPINE",
  "Story/game-led sequencing: the concept is DISCOVERED in play before it is named.",
  "Story -> definition -> diagram sequencing: the same experience hook, compressed, then promoted to formal treatment the same day."),
 ("5", "Visuals & Diagrams — what to draw, how to draw it, how to explain it, common mistakes. Simulation/Gamify.", "TIER SPLIT",
  "Informal 'sketches' and physical simulations — trade-off curves before formal PPCs.",
  "Exam-standard diagram discipline (labels, arrows, equilibria) — feeds directly into the adapter's diagram requirements."),
 ("6", "Worked Examples — numerical, graph-interpretation, and short case studies, reasoned step by step", "SPINE",
  "One concrete worked story-example.",
  "Numerical + graphical + case, each reasoned step by step with working shown."),
 ("7", "Practice Questions — Easy (concept check) -> Medium (application) -> Challenging (exam-style analysis)", "TIER SPLIT",
  "Easy tier is the core set; questions are game/mission-formatted.",
  "Medium/Challenging are the core set; every lesson closes with one exam-style question matching that lesson's drip-map row."),
 ("8", "Answers & Explanations — full solutions with WHY", "SPINE",
  "Full solutions, child-friendly language.",
  "Full solutions incl. mark-scheme-style annotations (what earns each mark)."),
 ("9", "EXAM ADAPTER SLOT (10-15 min) — the drip. Consult the 'Exam Skills Progression' sheet for THIS lesson's row: it names the exact IGCSE element and IB element to install here. Shared skills (definitions, diagrams, calculations) through M1; board-specific elements diverge from M2 L4.", "ADAPTER",
  "REPLACED by 'Think Like an Economist' — a 5-min reflection routine (What did we discover? Where else does this happen?). No exam content at Tier 1.",
  "The 10-15 min slot at lesson close. Content comes from the drip map — never invented per-lesson. IGCSE-bound students practise the gold-column element; IB-bound the green-column element; before M2 L4 both practise the shared element together."),
 ("10", "Homework — 10-15 mixed-difficulty questions incl. graph practice + one exam-style extended question", "TIER SPLIT",
  "Shorter set + one family-discussion mission.",
  "Full set incl. one extended question in the format of that lesson's adapter element."),
 ("11", "Extension Activities — article analysis, debate, mini-research, enrichment", "SPINE",
  "Game/mission-formatted versions of the same activities.",
  "Article analysis, debate, mini-research — the RWE (real-world example) bank starts collecting here from M3 onwards."),
 ("12", "Teaching Resources — glossary, summary sheet, formula/reference box, graph checklist, Key Takeaways, videos, readings, news sources", "SPINE",
  "Child-friendly glossary definitions.",
  "Exam-precise glossary definitions; graph checklist doubles as the adapter's diagram standard."),
]
r = 4
for num, section, layer, t1, t2 in trows:
    put(ws, r, [num, section, layer, t1, t2],
        fills=[None, None, LAYER_FILLS[layer], LIGHT_BLUE, LIGHT_GREEN])
    ws.cell(row=r, column=1).alignment = WRAP_C
    ws.cell(row=r, column=3).alignment = WRAP_C
    ws.cell(row=r, column=3).font = F(size=9, bold=True, color=LAYER_FONTS[layer])
    if layer == "ADAPTER":
        for col in range(1, 6):
            cc = ws.cell(row=r, column=col)
            cc.border = Border(left=thin, right=thin,
                               top=Side(style="medium", color=GOLD),
                               bottom=Side(style="medium", color=GOLD))
        ws.cell(row=r, column=2).font = F(size=10, bold=True)
    ws.row_dimensions[r].height = 66
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(row=r, column=1, value="Lesson 6 of each module additionally contains: Topic Review | Concept Connections | Common Misconceptions | Mock Exam (Tier 2 — format per the drip map: M1 L6 first levels-of-response/mini Part (b); M2 L6 data-response mock; M3 L6 full both-paper mock) or Capstone Simulation + Showcase Quiz (Tier 1) | Mark Scheme with model answers & examiner comments | Exam Strategy Guide. The three L6 lessons are the adapter's assembly checkpoints.")
c.font = F(size=9, italic=True, color=NAVY); c.alignment = WRAP
c.fill = PatternFill("solid", fgColor=CREAM)
ws.row_dimensions[r].height = 44
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
c = ws.cell(row=r, column=1, value="Maker's rule of thumb: editing a SPINE section = one edit serves both tiers. Editing a TIER SPLIT section = touch both delivery columns. Editing the ADAPTER slot = update the drip-map row first, then this template's slot follows it. Never write exam technique into sections 1-8.")
c.font = F(size=9, bold=True, color="7B4B2A"); c.alignment = WRAP
c.fill = PatternFill("solid", fgColor=LIGHT_GOLD)
ws.row_dimensions[r].height = 34

for col, w in zip("ABCDE", [4, 52, 12, 44, 50]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A4"

# ============================================================ Exam Skills Progression (Drip Map)
ws = wb.create_sheet("Exam Skills Progression")
title_block(ws, "Exam Adapter Drip Map — How skills are built inside the 18 lessons",
    "The adapter is not a separate module added after lesson 18. Each lesson's Exam-skill slot drip-feeds the adapter so it is fully assembled by M3 L6. Shared skills load first — board-specific skills diverge from M2 L4 onwards. Gold columns show which adapter element each lesson installs per board.", 6)

STAGE_COLS = {
    "Foundations": "2C4B6E", "Analysis": "3D6B58",
    "Evaluation":  "7B4B2A", "Data Response": "4B3D6E", "Synthesis": "5E4B28"
}
header_row(ws, 3,
    ["Stage", "Lesson", "Shared skill (both boards)", "Command terms", "IGCSE adapter element", "IB adapter element"],
    fills=[NAVY, NAVY, "4A7AB5", "4A7AB5", GOLD, GOLD])

drip = [
 ("Foundations","M1 L1","Precise definitions; PPC & circular flow drawn and labelled to exam standard","define, state, list, describe",
  "AO1: 1-sentence Cambridge command-word definitions; diagram labelling discipline",
  "AO1: 2-3 sentence IB-style full definitions; first 'explain' cause-chain"),
 ("Foundations","M1 L2","D/S diagram: shifts vs movements; schedules AND curves; equilibrium adjustment sequence","draw, label, outline, explain",
  "P1 MCQ training: single-best-answer discipline; diagram-reading in MCQ context",
  "Labelling rule: Price Level not P, Real Output not Q; IB 'explain' paragraph structure"),
 ("Analysis","M1 L3","PED/PES/YED calculations with full working; interpret the value; state the implication","calculate, explain, apply",
  "AO2: show all working; IGCSE 'calculate and comment' structure (2-3 marks)",
  "AO4: calculation technique; chain — value -> firm/government implication"),
 ("Analysis","M1 L4","Stakeholder analysis table; 2+ link cause-effect chains; first intervention diagrams","analyse, distinguish, apply",
  "AO2: structured analysis paragraphs; first Section B short-answer practice (4-6 marks)",
  "AO2: cause-effect-consequence chains; intro to 'examine' — two-sided, no final judgement yet"),
 ("Analysis","M1 L5","Cost/revenue calculations; advantages/disadvantages table; competitive vs monopoly comparison","calculate, compare, suggest",
  "AO2: numerical questions with working; competitive-vs-monopoly structured comparison",
  "AO2: apply theory to a firm scenario; 'analyse' paragraph — claim -> evidence -> link"),
 ("Evaluation","M1 L6 (Review)","First full evaluation paragraph: claim -> analysis -> counter -> judgement","discuss, evaluate, to what extent",
  "IGCSE levels-of-response intro: Band 1 vs Band 3 on a (d) question (8 marks); examiner's eye",
  "IB AO3 intro: 'discuss' structure — two sides + supported judgement; first mini Part (b)"),
 ("Evaluation","M2 L1","Labour-market D/S + minimum wage diagram; wage-difference structured analysis","draw, explain, analyse",
  "IGCSE: NMW diagram with surplus labelled; 'analyse the effect on workers' (4-mark answer)",
  "IB: policy chain — NMW -> labour market effects; link to equity key concept"),
 ("Evaluation","M2 L2","GDP/GNI calculations; real vs nominal; growth rate; per capita — all with working","calculate, construct, comment",
  "IGCSE AO2: GDP numerical questions; limitations of GDP as a 3-mark structured list",
  "IB AO4: calculations with units; AO3: 'to what extent is GDP a good measure of living standards?'"),
 ("Evaluation","M2 L3","AD/AS diagrams: full labelling, both AS shapes, shift analysis, output-gap identification","draw, label, explain, analyse",
  "IGCSE: NO AD-AS diagram exists in 0455 \u2014 practise verbal 'effect of X on total demand / output' chains instead (3-4 mark structure); treat the model as scaffolding",
  "IB: both AS shapes drawn and defended; 'examine' Keynesian vs monetarist view — two-sided"),
 ("Data Response","M2 L4","Unemployment rate & CPI inflation calculations from raw data; types with policy link","calculate, comment, identify",
  "IGCSE P2 Section A (data response): extract data, calculate, comment — Q1 format. First timed Section A practice",
  "IB P2 data-response intro: read extract -> define -> calculate -> analyse in sub-questions (a)-(f)"),
 ("Data Response","M2 L5","Policy evaluation written to exam length: mechanism -> strengths -> limitations -> judgement","discuss, evaluate, recommend, to what extent",
  "IGCSE (d)-question full practice (8 marks): two-sided + supported conclusion using levels-of-response criteria",
  "IB AO3: 'to what extent is fiscal/monetary policy effective?' — two sides, real-world case, conditioned judgement"),
 ("Data Response","M2 L6 (Review)","Full data-response mock under timed conditions; self-mark against a mark scheme","all above",
  "IGCSE P2 Section A mock: Q1 all parts (a)-(f), 20 marks, 40 min. Mark-scheme literacy debrief",
  "IB P2 mock: structured data-response, 20 marks. Technique: quote data as evidence — never assert without extracting"),
 ("Synthesis","M3 L1","Multi-step numerical reasoning; structured explanation writing: each step earns a mark","explain, calculate, justify",
  "IGCSE: 3-4 step structured explanation — signal words: 'therefore', 'this means that'",
  "IB: first real-world example used as evidence (not decoration) — name it, state the fact, link to the concept"),
 ("Synthesis","M3 L2","Balanced argument: tool -> who wins -> who loses -> net judgement, exam-length","examine, discuss, weigh",
  "IGCSE (d)-question on protectionism (8 marks): independent full attempt, self-mark",
  "IB Part (a) practice: 'Examine the effects of a tariff' — 10 marks, 15 min; one side per paragraph, explicit link words"),
 ("Synthesis","M3 L3","Currency diagrams × 4 cases; cross-currency calculations; BoP components + calculations","draw, calculate, analyse",
  "IGCSE: currency diagram + price-conversion calculation in one data-response question; BoP calculation",
  "IB: four-diagram sequence — cause -> exchange rate -> export price -> AD effect; BoP beyond current account"),
 ("Synthesis","M3 L4","Lorenz curve interpretation; poverty-policy evaluation; population data analysis","interpret, evaluate, recommend",
  "IGCSE: living-standards comparison using HDI & GDP data (4-6 marks); policy-to-reduce-poverty structured answer",
  "IB 3.4: Gini comparison + 'evaluate the effectiveness of progressive taxation in reducing inequality' (AO3 full paragraph)"),
 ("Synthesis","M3 L5","Development strategy comparison; justify/recommend with real-world evidence","compare, contrast, justify, recommend",
  "IGCSE Section B extended response (12-16 marks): full independent attempt on development question + self-mark",
  "IB P1 Part (b) full practice: 'Evaluate whether free trade promotes economic development' — 15 marks, 20 min; RWE required"),
 ("Synthesis","M3 L6 (Review)","Full mock examination — both paper formats — under exam conditions","all command terms",
  "IGCSE full mock: P1 (40 MCQ, 60 min) + P2 Section A (20 marks) + Section B (3 of 4, 60 marks). Timing strategy + mark-scheme debrief",
  "IB full mock: P1 two essays (a)+(b) × 2 questions; P2 data response; P3 HL policy paper. RWE bank review; essay structure final check"),
]

r = 4
current_stage = None
stage_start = 4
for row in drip:
    stage, lesson, shared, cmds, igcse_el, ib_el = row
    if stage != current_stage:
        if current_stage is not None:
            # style all cells in the completed band
            for rr in range(stage_start, r):
                ws.cell(row=rr, column=1).fill = PatternFill("solid", fgColor=STAGE_COLS[current_stage])
            ws.merge_cells(start_row=stage_start, start_column=1, end_row=r-1, end_column=1)
            ca = ws.cell(row=stage_start, column=1)
            ca.value = current_stage; ca.font = F(size=10, bold=True, color="FFFFFF")
            ca.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True, text_rotation=90)
        current_stage = stage; stage_start = r
    review = "(Review)" in lesson
    bf = [None, LIGHT_GOLD if review else None, LIGHT_BLUE, None, LIGHT_GOLD, LIGHT_GREEN]
    put(ws, r, ["", lesson, shared, cmds, igcse_el, ib_el], fills=bf)
    ws.cell(row=r, column=2).font = F(size=10, bold=True)
    ws.row_dimensions[r].height = 78; r += 1

# close final stage band
for rr in range(stage_start, r):
    ws.cell(row=rr, column=1).fill = PatternFill("solid", fgColor=STAGE_COLS[current_stage])
ws.merge_cells(start_row=stage_start, start_column=1, end_row=r-1, end_column=1)
ca = ws.cell(row=stage_start, column=1)
ca.value = current_stage; ca.font = F(size=10, bold=True, color="FFFFFF")
ca.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True, text_rotation=90)

# footer
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws.cell(row=r, column=1, value="IB AOs — AO1 Knowledge & understanding · AO2 Application & analysis · AO3 Synthesis & evaluation · AO4 Skills & techniques. IGCSE 0455 (2027-29) AOs — AO1 Knowledge 43% · AO2 Analysis 47% · AO3 Evaluation 10%. Blue = shared skill | Gold = IGCSE adapter element | Green = IB adapter element | Gold row = Review lesson.")
c.font = F(size=9, italic=True, color=NAVY); c.alignment = WRAP
c.fill = PatternFill("solid", fgColor=CREAM); ws.row_dimensions[r].height = 34

for col, w in zip("ABCDEF", [5, 15, 46, 26, 52, 52]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "C4"
ws.auto_filter.ref = f"B3:F{r-2}"

# ============================================================ Assessment Styles
ws = wb.create_sheet("Assessment Styles")
title_block(ws, "Assessment Styles Practised (Pre-IB track)", "All questions original — no copyrighted past-paper items are reproduced. Pre-IGCSE track assesses through games, simulations and short friendly checks instead.", 3)
header_row(ws, 3, ["Question type", "Modelled on", "First appears"])
arows = [
 ("Multiple choice", "IGCSE 0455 Paper 1", "M1 L1 (quick checks) -> full set M1 L6"),
 ("Short structured", "IGCSE 0455 Paper 2 / IB Paper 2 (a)-(f)", "M1 L2"),
 ("Calculations (show working)", "IGCSE numerical (AO2) / IB Paper 3 (HL policy paper — calculations + policy recommendation)", "M1 L3"),
 ("Data response", "IB Paper 2", "M2 L3"),
 ("Extended response (a + b)", "IB Paper 1", "M1 L4 (scaffolded) -> full M3 L6"),
]
r = 4
for row in arows:
    put(ws, r, list(row), bold_first=True); r += 1
for col, w in zip("ABC", [26, 60, 42]):
    ws.column_dimensions[col].width = w


# ============================================================ Coverage Audit
OK="[covered]"; PART="[partial]"; MISS="[gap]"; OVER="[beyond syllabus]"
def audit_sheet(name, title, subtitle, rows):
    ws = wb.create_sheet(name)
    title_block(ws, title, subtitle, 4)
    header_row(ws, 3, ["Syllabus reference", "Status", "Where in this course map", "Notes / action"])
    r = 4
    for ref, st, where, note in rows:
        fill = None
        if st.startswith("[gap]"): fill = "F8D7DA"
        elif st.startswith("[partial]"): fill = FIX
        elif st.startswith("[beyond syllabus]"): fill = LIGHT_GOLD
        put(ws, r, [ref, st, where, note], fills=[None, fill, None, fill])
        ws.cell(row=r, column=1).font = F(size=10, bold=True)
        r += 1
    for col, w in zip("ABCD", [46, 16, 26, 78]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:D{r-1}"
    return r

ig = [
 ("1.1 The nature of the basic economic problem", OK, "M1 L1", "Incl. economic vs free goods (now explicit in lesson content)."),
 ("1.2 Factors of production (+ rewards)", OK, "M1 L1", "Four rewards (rent, wages, interest, profit) now named in lesson content."),
 ("1.3 Opportunity cost", OK, "M1 L1", ""),
 ("1.4 PPC diagrams", OK, "M1 L1", "Points under/on/beyond, movements, shifts — all present."),
 ("2.1 Role of markets in allocating resources", OK, "M1 L2", ""),
 ("2.2 Demand", OK, "M1 L2", ""),
 ("2.3 Supply", OK, "M1 L2", ""),
 ("2.4 Price determination (equilibrium/disequilibrium)", OK, "M1 L2", "Schedules (tables) AND curves now both named — IGCSE tests both."),
 ("2.5 Price changes (causes & consequences)", OK, "M1 L2", ""),
 ("2.6 PED", OK, "M1 L3", "Full five-value vocabulary (perfectly inelastic -> perfectly elastic) now named."),
 ("2.7 PES", OK, "M1 L3", ""),
 ("2.8 Market economic system (for & against)", OK, "M1 L1", "Explicit for/against now in lesson content."),
 ("2.9 Market failure", OK, "M1 L4", "Monopoly abuse + non-provision of public goods now included. Diagrams NOT required at IGCSE (flagged in-sheet)."),
 ("2.10 Mixed economic system & intervention", OK, "M1 L1 + L4", "CLOSED in v3: regulation, privatisation, nationalisation, direct provision, quotas + for/against the mixed system added to L4."),
 ("3.1 Money and banking", OK, "M2 L1 (NEW)", "CLOSED in v3: forms/functions/characteristics of money; central & commercial banks."),
 ("3.2 Households (spending, saving, borrowing)", OK, "M2 L1 (NEW)", "CLOSED in v3: income, interest rate, confidence, age, culture."),
 ("3.3 Workers (wages, labour market, NMW, mobility, division of labour)", OK, "M2 L1 (NEW)", "CLOSED in v3: labour-market D/S + NMW diagrams to exam standard."),
 ("3.4 Firms (types, mergers, economies of scale)", OK, "M1 L5 (NEW)", "CLOSED in v3: merger types + ATC diagram for economies/diseconomies of scale."),
 ("3.5 Firms and production", OK, "M1 L5 (NEW)", "CLOSED in v3: labour- vs capital-intensive; production vs productivity; investment effects."),
 ("3.6 Firms' costs, revenue and objectives", OK, "M1 L5 (NEW)", "CLOSED in v3: TC/ATC/FC/AFC/VC/AVC and TR/AR calculations; four objectives."),
 ("3.7 Types of markets (competitive vs monopoly)", OK, "M1 L5 (NEW)", "CLOSED in v3: characteristics & effects on price/quality/choice/profit — no diagrams required."),
 ("4.1 Government macroeconomic intervention (aims & conflicts)", OK, "M2 L2 + L5", "All six aims now named in M2 L5, with the three IGCSE conflict pairs."),
 ("4.2 Fiscal policy (budget, spending, taxation)", OK, "M2 L5", "Budget deficit/surplus calculations + tax classifications + reasons for taxation now explicit."),
 ("4.3 Monetary policy", OK, "M2 L5", "Interest rate + money supply + exchange-rate channel all named."),
 ("4.4 Supply-side policy", OK, "M2 L5", "Infrastructure spending included."),
 ("4.5 Economic growth (+ recession)", OK, "M2 L2 + L4", "Recession causes & consequences explicit; PPC framing labelled IB-style."),
 ("4.6 Employment and unemployment", OK, "M2 L4", "Labour force survey; seasonal unemployment explicit."),
 ("4.7 Inflation", OK, "M2 L4", "Deflation flagged as IB-only (removed from IGCSE 2027-29)."),
 ("5.1 Living standards (GDP per head, HDI)", OK, "M2 L2 + M3 L4-L5", "Distribution comparisons now reinforced by the new inequality lesson."),
 ("5.2 Poverty (absolute/relative, causes, policies)", OK, "M3 L4 (NEW)", "CLOSED in v3 — and now anchors the Equity key concept."),
 ("5.3 Population (birth/death rates, migration, structure)", OK, "M3 L4 (NEW)", "CLOSED in v3: optimum population; size & structure effects. (Population pyramids removed in 2027-29 — not taught as examinable.)"),
 ("5.4 Differences in economic development", OK, "M3 L5", ""),
 ("6.1 Specialisation and free trade", OK, "M3 L1", "Formal comparative advantage flagged as IB HL."),
 ("6.2 Globalisation and trade restrictions", OK, "M3 L2 + L5", "CLOSED in v3: MNCs (host/home) in L5; embargoes + full reasons-for-restrictions list in L2."),
 ("6.3 Foreign exchange rates", OK, "M3 L3", "Floating only at IGCSE (fixed = IB HL, flagged); reasons for buying/selling currencies incl. remittances now explicit."),
 ("6.4 Current account of the balance of payments", OK, "M3 L3", "Four components + deficit/surplus calculations + stability policies explicit."),
]
audit_sheet("IGCSE 0455 Audit", "Coverage Audit — Cambridge IGCSE 0455 (2027-2029), all 35 subtopics — FULLY CLOSED in v3",
    "Status: [covered]. Every subtopic now has a home lesson; 'CLOSED in v3' marks items resolved by the three new lessons or by named additions to existing lessons.", ig)

ib = [
 ("1.1 What is economics?", OK, "M1 L1", "Economics as a social science + positive vs normative now explicit alongside scarcity/choice/PPC/circular flow."),
 ("1.2 How do economists approach the world?", OK, "M1 L1", "CLOSED in v3: methodology & pluralism folded into L1 (with the 'fact or opinion' routine on the Pre-IGCSE track)."),
 ("2.1 Demand", OK, "M1 L2", ""),
 ("2.2 Supply", OK, "M1 L2", ""),
 ("2.3 Competitive market equilibrium", OK, "M1 L2", ""),
 ("2.4 Critique of maximising behaviour — behavioural economics (HL)", OK, "M1 L2 (enrichment)", "CLOSED in v3 as HL enrichment: biases & nudges segment + Pre-IGCSE 'trick your brain' experiment."),
 ("2.5 Elasticities of demand (PED, YED)", OK, "M1 L3", "XED remains flagged (!) as beyond both current syllabi."),
 ("2.6 Elasticity of supply (PES)", OK, "M1 L3", ""),
 ("2.7 Role of government in microeconomics", OK, "M1 L4", "Direct provision + regulation now named alongside price controls, taxes, subsidies."),
 ("2.8 Market failure — externalities & common pool resources", OK, "M1 L4", "CLOSED in v3: common pool resources + tragedy of the commons (fishing-lake game) added."),
 ("2.9 Market failure — public goods", OK, "M1 L4", ""),
 ("2.10 Market failure — asymmetric information (HL)", OK, "M1 L5 (enrichment)", "CLOSED in v3 as HL enrichment: 'lemons' used-bikes mini-game."),
 ("2.11 Market failure — market power (HL)", OK, "M1 L5 (NEW)", "CLOSED in v3: competitive vs monopoly + HL market-structure preview (also serves IGCSE 3.7)."),
 ("2.12 The market's inability to achieve equity (HL)", OK, "M3 L4 (NEW)", "CLOSED in v3: paired with the inequality lesson as an HL preview."),
 ("3.1 Measuring economic activity", OK, "M2 L2", ""),
 ("3.2 Variations in economic activity — AD/AS", OK, "M2 L3", ""),
 ("3.3 Macroeconomic objectives", OK, "M2 L2 + L4-L5", "HL 'sustainable level of government debt' now an enrichment line in M2 L5."),
 ("3.4 Economics of inequality and poverty (SL CORE)", OK, "M3 L4 (NEW)", "CLOSED in v3: Lorenz curve, Gini coefficient, poverty measurement, redistribution — the former biggest IB gap, now the Equity anchor lesson."),
 ("3.5 Demand management — monetary policy", OK, "M2 L5", "Strengthened by the new money & banking opener (M2 L1)."),
 ("3.6 Demand management — fiscal policy", OK, "M2 L5", "HL Keynesian multiplier flagged as enrichment."),
 ("3.7 Supply-side policies", OK, "M2 L5", ""),
 ("4.1 Benefits of international trade", OK, "M3 L1", "Comparative advantage = HL only (flagged)."),
 ("4.2 Types of trade protection", OK, "M3 L2", ""),
 ("4.3 Arguments for and against trade control", OK, "M3 L2", ""),
 ("4.4 Economic integration", OK, "M3 L5", "CLOSED in v3: integration ladder named (preferential -> FTA -> customs union -> common market -> monetary union; depth = HL)."),
 ("4.5 Exchange rates", OK, "M3 L3", "Fixed/managed = HL (flagged)."),
 ("4.6 Balance of payments", OK, "M3 L3", "CLOSED in v3: financial account added as the IB layer beyond the IGCSE current-account scope."),
 ("4.7 Sustainable development", OK, "M3 L5", ""),
 ("4.8 Measuring development", OK, "M3 L5", ""),
 ("4.9 Barriers to economic growth and development", OK, "M3 L5", ""),
 ("4.10 Economic growth and development strategies", OK, "M3 L5", "CLOSED in v3: strategies named (trade strategies, diversification, market-based vs interventionist, aid, FDI, institutional change)."),
]
audit_sheet("IB DP Audit", "Coverage Audit — IB DP Economics (2022 guide, current), Units 1-4 — FULLY CLOSED in v3",
    "Status: [covered]. SL core is fully covered; HL-only items are covered as flagged enrichment, appropriate for a foundation course.", ib)

# ---- Recommendation sheet
ws = wb.create_sheet("Coverage Summary")
title_block(ws, "Coverage Summary — v3 (18 lessons): audit fully closed",
    "Verdict: with the three new lessons and the named in-lesson additions, every Cambridge IGCSE 0455 (2027-2029) subtopic and every IB DP Economics (2022 guide) SL-core item now has a home lesson; HL items are covered as flagged enrichment.", 2)
rows = [
 ("What changed in v3", "Three lessons added (15 -> 18, modules now 6 lessons each, review moves to Lesson 6): M1 L5 'Firms, Costs & Market Power' (IGCSE 3.4-3.7 + IB HL 2.11/2.10 enrichment); M2 L1 'Money, Banking, Households & Workers' as the macro opener (IGCSE 3.1-3.3 — money bridges naturally into monetary policy in M2 L5); M3 L4 'Inequality, Poverty & Population' before development (IGCSE 5.2-5.3 + IB 3.4 SL core, Lorenz/Gini)."),
 ("Key-concept fix", "Equity — previously the only key concept without a content anchor despite being marked PP in the Key-Concept Map — is now anchored by M3 L4 (and reinforced by the tax-fairness segment in M2 L5)."),
 ("Minor additions folded into existing lessons", "Positive vs normative + economics as a social science (M1 L1); behavioural-economics enrichment (M1 L2); monopoly as market failure, common pool resources, regulation/privatisation/nationalisation/direct provision/quotas (M1 L4); six macro aims by name, tax classifications, budget calculations, HL government-debt line (M2 L5); recession + seasonal unemployment explicit (M2 L4); embargoes + full reasons-for-restrictions (M3 L2); remittances etc. + financial account (M3 L3); MNCs + integration ladder + named development strategies (M3 L5)."),
 ("Over-teach flags retained ((!) in module sheets)", "XED (in neither current syllabus); deflation (IB only); fixed/managed exchange rates (IB HL only); formal comparative advantage (IB HL only). These stay visible so tutors treat them as enrichment, not core."),
 ("Remaining judgement calls (deliberate, not gaps)", "Population pyramids (removed from IGCSE 2027-29) are not taught as examinable; claimant count (removed) not taught; IB internal assessment & Paper-specific drilling belong to the actual IB course, not this foundation bridge."),
 ("Timing note", "18 lessons × 90-120 min. The three new lessons are content-dense (esp. M2 L1); if a cohort runs short, the HL enrichment segments are the designed release valve — cut those before cutting board-required content."),
]
r = 4
for k, v in rows:
    put(ws, r, [k, v], bold_first=True)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=GREY)
    ws.row_dimensions[r].height = 68
    r += 1
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 130

# Reorder: place Lesson Template immediately after Key-Concept Map
desired = ["IGCSE to IB Bridge", "Overview", "Verification Notes", "Structure",
           "Key-Concept Map", "Lesson Template",
           "M1 Microeconomics", "M2 Macroeconomics", "M3 Global Economics",
           "Exam Skills Progression", "Assessment Styles",
           "IGCSE 0455 Audit", "IB DP Audit", "Coverage Summary"]
present = [s for s in desired if s in wb.sheetnames] + [s for s in wb.sheetnames if s not in desired]
wb._sheets.sort(key=lambda ws: present.index(ws.title))

out = "outputs/Economics-Academy-Course-Map-v3.xlsx"
wb.save(out)
print("saved", out)
